import pandas as pd
from openpyxl import load_workbook

excel_file = 'Documents/Input/toby.xlsx'
df = pd.read_excel(excel_file)

# *********************** carriageway_toby loc no and wayleave no  ************************************************

carriageway_toby = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False)]

index = carriageway_toby.index
num_of_rows_cw_toby = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B31"] = num_of_rows_cw_toby
wb.save(filename)

# *********************** footway_toby loc no and wayleave no  ************************************************

footway_toby = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False)]

index = footway_toby.index
num_of_rows_fw_toby = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B32"] = num_of_rows_fw_toby
wb.save(filename)

# *********************** grass verge_fw2 loc no and wayleave no  ************************************************

grassverge_toby = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == False)]

index = grassverge_toby.index
num_of_rows_verge_toby = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B33"] = num_of_rows_verge_toby
wb.save(filename)

# *********************** carriageway_toby loc no and wayleave yes  ************************************************

carriageway_toby_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True)]

index = carriageway_toby_way.index
num_of_rows_cw_toby_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D31"] = num_of_rows_cw_toby_way
wb.save(filename)

# *********************** footway_toby loc no and wayleave yes  ************************************************

footway_toby_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True)]

index = footway_toby_way.index
num_of_rows_fw_toby_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D32"] = num_of_rows_fw_toby_way
wb.save(filename)

# *********************** grass verge_toby loc no and wayleave yes  ************************************************

grassverge_toby_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == True)]

index = grassverge_toby_way.index
num_of_rows_verge_toby_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D33"] = num_of_rows_verge_toby_way
wb.save(filename)

# *********************** carriageway_toby loc yes  ************************************************

carriageway_toby_loc = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == True)]

index = carriageway_toby_loc.index
num_of_rows_cw_toby_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C31"] = num_of_rows_cw_toby_loc
wb.save(filename)

# *********************** footway_toby loc yes  ************************************************

footway_toby_loc = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == True)]

index = footway_toby_loc.index
num_of_rows_fw_toby_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C32"] = num_of_rows_fw_toby_loc
wb.save(filename)

# *********************** grass verge_toby loc yes  ************************************************

grassverge_toby_loc = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Grass Verge') & (df['loc'] == True)]

index = grassverge_toby_loc.index
num_of_rows_verge_toby_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C33"] = num_of_rows_verge_toby_loc
wb.save(filename)