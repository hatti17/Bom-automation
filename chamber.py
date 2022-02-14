import pandas as pd
from openpyxl import load_workbook

excel_file = 'Documents/Input/chamber.xlsx'
df = pd.read_excel(excel_file)

# *********************** carriageway_fw2 loc no and wayleave no  ************************************************

carriageway_fw2 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'FW2')]

index = carriageway_fw2.index
num_of_rows_cw_fw2 = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B13"] = num_of_rows_cw_fw2
wb.save(filename)

# *********************** footway_fw2 loc no and wayleave no  ************************************************

footway_fw2 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'FW2')]

index = footway_fw2.index
num_of_rows_fw_fw2 = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C13"] = num_of_rows_fw_fw2
wb.save(filename)

# *********************** grass verge_fw2 loc no and wayleave no  ************************************************

grassverge_fw2 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'FW2')]

index = grassverge_fw2.index
num_of_rows_verge_fw2 = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D13"] = num_of_rows_verge_fw2
wb.save(filename)

# *********************** carriageway_fw4 loc no and wayleave no  ************************************************

carriageway_fw4 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'FW4')]

index = carriageway_fw4.index
num_of_rows_cw_fw4 = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B15"] = num_of_rows_cw_fw4
wb.save(filename)

# *********************** footway_fw4 loc no and wayleave no  ************************************************

footway_fw4 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'FW4')]

index = footway_fw4.index
num_of_rows_fw_fw4 = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C15"] = num_of_rows_fw_fw4
wb.save(filename)

# *********************** grass verge_fw4 loc no and wayleave no  ************************************************

grassverge_fw4 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'FW4')]

index = grassverge_fw4.index
num_of_rows_verge_fw4 = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D15"] = num_of_rows_verge_fw4
wb.save(filename)

# *********************** carriageway_fw6 loc no and wayleave no  ************************************************

carriageway_fw6 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'FW6')]

index = carriageway_fw6.index
num_of_rows_cw_fw6 = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B16"] = num_of_rows_cw_fw6
wb.save(filename)

# *********************** footway_fw6 loc no and wayleave no  ************************************************

footway_fw6 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'FW6')]

index = footway_fw6.index
num_of_rows_fw_fw6 = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C16"] = num_of_rows_fw_fw6
wb.save(filename)

# *********************** grass verge_fw6 loc no and wayleave no  ************************************************

grassverge_fw6 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'FW6')]

index = grassverge_fw6.index
num_of_rows_verge_fw6 = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D16"] = num_of_rows_verge_fw6
wb.save(filename)


# *********************** carriageway_fw2 loc no and wayleave yes  ************************************************

carriageway_fw2_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'FW2')]

index = carriageway_fw2_way.index
num_of_rows_cw_fw2_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["H13"] = num_of_rows_cw_fw2_way
wb.save(filename)

# *********************** footway_fw2 loc no and wayleave yes  ************************************************

footway_fw2_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'FW2')]

index = footway_fw2_way.index
num_of_rows_fw_fw2_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["I13"] = num_of_rows_fw_fw2_way
wb.save(filename)

# *********************** grass verge_fw2 loc no and wayleave yes  ************************************************

grassverge_fw2_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'FW2')]

index = grassverge_fw2_way.index
num_of_rows_verge_fw2_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J13"] = num_of_rows_verge_fw2_way
wb.save(filename)

# *********************** carriageway_fw4 loc no and wayleave yes  ************************************************

carriageway_fw4_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'FW4')]

index = carriageway_fw4_way.index
num_of_rows_cw_fw4_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["H15"] = num_of_rows_cw_fw4_way
wb.save(filename)

# *********************** footway_fw4 loc no and wayleave yes  ************************************************

footway_fw4_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'FW4')]

index = footway_fw4_way.index
num_of_rows_fw_fw4_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["I15"] = num_of_rows_fw_fw4_way
wb.save(filename)

# *********************** grass verge_fw4 loc no and wayleave yes  ************************************************

grassverge_fw4_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'FW4')]

index = grassverge_fw4_way.index
num_of_rows_verge_fw4_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J15"] = num_of_rows_verge_fw4_way
wb.save(filename)

# *********************** carriageway_fw6 loc no and wayleave yes  ************************************************

carriageway_fw6_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'FW6')]

index = carriageway_fw6_way.index
num_of_rows_cw_fw6_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["H16"] = num_of_rows_cw_fw6_way
wb.save(filename)

# *********************** footway_fw6 loc no and wayleave yes  ************************************************

footway_fw6_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'FW6')]

index = footway_fw6_way.index
num_of_rows_fw_fw6_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["I16"] = num_of_rows_fw_fw6_way
wb.save(filename)

# *********************** grass verge_fw6 loc no and wayleave yes  ************************************************

grassverge_fw6_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'FW6')]

index = grassverge_fw6_way.index
num_of_rows_verge_fw6_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J16"] = num_of_rows_verge_fw6_way
wb.save(filename)

# *********************** carriageway_fw2 loc yes  ************************************************

carriageway_fw2_loc = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == True)
                         & (df['type'] == 'FW2')]

index = carriageway_fw2_loc.index
num_of_rows_cw_fw2_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E13"] = num_of_rows_cw_fw2_loc
wb.save(filename)

# *********************** footway_fw2 loc yes  ************************************************

footway_fw2_loc = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == True)
                         & (df['type'] == 'FW2')]

index = footway_fw2_loc.index
num_of_rows_fw_fw2_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["F13"] = num_of_rows_fw_fw2_loc
wb.save(filename)

# *********************** grass verge_fw2 loc yes  ************************************************

grassverge_fw2_loc = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Grass Verge') & (df['loc'] == True)
                         & (df['type'] == 'FW2')]

index = grassverge_fw2_loc.index
num_of_rows_verge_fw2_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["G13"] = num_of_rows_verge_fw2_loc
wb.save(filename)

# *********************** carriageway_fw4 loc yes  ************************************************

carriageway_fw4_loc = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == True)
                         & (df['type'] == 'FW4')]

index = carriageway_fw4_loc.index
num_of_rows_cw_fw4_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E15"] = num_of_rows_cw_fw4_loc
wb.save(filename)

# *********************** footway_fw4 loc yes  ************************************************

footway_fw4_loc = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == True)
                         & (df['type'] == 'FW4')]

index = footway_fw4_loc.index
num_of_rows_fw_fw4_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["F15"] = num_of_rows_fw_fw4_loc
wb.save(filename)

# *********************** grass verge_fw4 loc yes  ************************************************

grassverge_fw4_loc = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Grass Verge') & (df['loc'] == True)
                         & (df['type'] == 'FW4')]

index = grassverge_fw4_loc.index
num_of_rows_verge_fw4_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["G15"] = num_of_rows_verge_fw4_loc
wb.save(filename)

# *********************** carriageway_fw6 loc yes  ************************************************

carriageway_fw6_loc = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == True)
                         & (df['type'] == 'FW6')]

index = carriageway_fw6_loc.index
num_of_rows_cw_fw6_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E16"] = num_of_rows_cw_fw6_loc
wb.save(filename)

# *********************** footway_fw6 loc yes  ************************************************

footway_fw6_loc = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == True)
                         & (df['type'] == 'FW6')]

index = footway_fw6_loc.index
num_of_rows_fw_fw6_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["F16"] = num_of_rows_fw_fw6_loc
wb.save(filename)

# *********************** grass verge_fw6 loc yes  ************************************************

grassverge_fw6_loc = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Grass Verge') & (df['loc'] == True)
                         & (df['type'] == 'FW6')]

index = grassverge_fw6_loc.index
num_of_rows_verge_fw6_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["G16"] = num_of_rows_verge_fw6_loc
wb.save(filename)

# *********************** carriageway_JB loc no and wayleave no  ************************************************

carriageway_JB = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'JB')]

index = carriageway_JB.index
num_of_rows_cw_JB = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K18"] = num_of_rows_cw_JB
wb.save(filename)

# *********************** footway_JB loc no and wayleave no  ************************************************

footway_JB = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'JB')]

index = footway_JB.index
num_of_rows_fw_JB = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L18"] = num_of_rows_fw_JB
wb.save(filename)

# *********************** grass verge_JB loc no and wayleave no  ************************************************

grassverge_JB = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'JB')]

index = grassverge_JB.index
num_of_rows_verge_JB = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M18"] = num_of_rows_verge_JB
wb.save(filename)

# *********************** carriageway_JB loc no and wayleave yes  ************************************************

carriageway_JB_way = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'JB')]

index = carriageway_JB_way.index
num_of_rows_cw_JB_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["Q18"] = num_of_rows_cw_JB_way
wb.save(filename)

# *********************** footway_JB loc no and wayleave yes  ************************************************

footway_JB_way = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'JB')]

index = footway_JB_way.index
num_of_rows_fw_JB_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["R18"] = num_of_rows_fw_JB_way
wb.save(filename)

# *********************** grass verge_JB loc no and wayleave yes  ************************************************

grassverge_JB_way = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'JB')]

index = grassverge_JB_way.index
num_of_rows_verge_JB_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["S18"] = num_of_rows_verge_JB_way
wb.save(filename)

# *********************** carriageway_JB loc yes  ************************************************

carriageway_JB_loc = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Carriageway') & (df['loc'] == True)
                         & (df['type'] == 'JB')]

index = carriageway_JB_loc.index
num_of_rows_cw_JB_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["N18"] = num_of_rows_cw_JB_loc
wb.save(filename)

# *********************** footway_JB loc yes  ************************************************

footway_JB_loc = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Footway') & (df['loc'] == True)
                         & (df['type'] == 'JB')]

index = footway_JB_loc.index
num_of_rows_fw_JB_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["O18"] = num_of_rows_fw_JB_loc
wb.save(filename)

# *********************** grass verge_JB loc yes  ************************************************

grassverge_JB_loc = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Grass Verge') & (df['loc'] == True)
                         & (df['type'] == 'JB')]

index = grassverge_JB_loc.index
num_of_rows_verge_JB_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["P18"] = num_of_rows_verge_JB_loc
wb.save(filename)

# *********************** carriageway_MH loc no and wayleave no  ************************************************

carriageway_MH = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'ManHole')]

index = carriageway_MH.index
num_of_rows_cw_MH = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K19"] = num_of_rows_cw_MH
wb.save(filename)

# *********************** footway_MH loc no and wayleave no  ************************************************

footway_MH = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'ManHole')]

index = footway_MH.index
num_of_rows_fw_MH = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L19"] = num_of_rows_fw_MH
wb.save(filename)

# *********************** grass verge_MH loc no and wayleave no  ************************************************

grassverge_MH = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'ManHole')]

index = grassverge_MH.index
num_of_rows_verge_MH = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M19"] = num_of_rows_verge_MH
wb.save(filename)

# *********************** carriageway_MH loc no and wayleave yes  ************************************************

carriageway_MH_way = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'ManHole')]

index = carriageway_MH_way.index
num_of_rows_cw_MH_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["Q19"] = num_of_rows_cw_MH_way
wb.save(filename)

# *********************** footway_MH loc no and wayleave yes  ************************************************

footway_MH_way = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'ManHole')]

index = footway_MH_way.index
num_of_rows_fw_MH_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["R19"] = num_of_rows_fw_MH_way
wb.save(filename)

# *********************** grass verge_MH loc no and wayleave yes  ************************************************

grassverge_MH_way = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'ManHole')]

index = grassverge_MH_way.index
num_of_rows_verge_MH_way = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["S19"] = num_of_rows_verge_MH_way
wb.save(filename)

# *********************** carriageway_MH loc yes  ************************************************

carriageway_MH_loc = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Carriageway') & (df['loc'] == True)
                         & (df['type'] == 'ManHole')]

index = carriageway_MH_loc.index
num_of_rows_cw_MH_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["N19"] = num_of_rows_cw_MH_loc
wb.save(filename)

# *********************** footway_MH loc yes  ************************************************

footway_MH_loc = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Footway') & (df['loc'] == True)
                         & (df['type'] == 'ManHole')]

index = footway_MH_loc.index
num_of_rows_fw_MH_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["O19"] = num_of_rows_fw_MH_loc
wb.save(filename)

# *********************** grass verge_MH loc yes  ************************************************

grassverge_MH_loc = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Grass Verge') & (df['loc'] == True)
                         & (df['type'] == 'ManHole')]

index = grassverge_MH_loc.index
num_of_rows_verge_MH_loc = len(index)
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["P19"] = num_of_rows_verge_MH_loc
wb.save(filename)

