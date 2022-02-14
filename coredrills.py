import pandas as pd
from openpyxl import load_workbook

excel_file = 'Documents/Input/chamber.xlsx'
df = pd.read_excel(excel_file)

# *********************** carriageway_CD loc no and wayleave no  ************************************************

carriageway_CD = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'JB')]

index = carriageway_CD.index
num_of_rows_cw_CD = carriageway_CD['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B26"] = num_of_rows_cw_CD
wb.save(filename)

# *********************** footway_CD loc no and wayleave no  ************************************************

footway_CD = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'JB')]

index = footway_CD.index
num_of_rows_fw_CD = footway_CD['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C26"] = num_of_rows_fw_CD
wb.save(filename)

# *********************** grass verge_CD loc no and wayleave no  ************************************************

grassverge_CD = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'JB')]

index = grassverge_CD.index
num_of_rows_verge_CD = grassverge_CD['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D26"] = num_of_rows_verge_CD
wb.save(filename)

# *********************** carriageway_CD loc no and wayleave yes  ************************************************

carriageway_CD_way = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'JB')]

index = carriageway_CD_way.index
num_of_rows_cw_CD_way = carriageway_CD_way['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["H26"] = num_of_rows_cw_CD_way
wb.save(filename)

# *********************** footway_CD loc no and wayleave yes  ************************************************

footway_CD_way = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'JB')]

index = footway_CD_way.index
num_of_rows_fw_CD_way = footway_CD_way['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["I26"] = num_of_rows_fw_CD_way
wb.save(filename)

# *********************** grass verge_CD loc no and wayleave yes  ************************************************

grassverge_CD_way = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'JB')]

index = grassverge_CD_way.index
num_of_rows_verge_CD_way = grassverge_CD_way['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J26"] = num_of_rows_verge_CD_way
wb.save(filename)

# *********************** carriageway_CD loc yes  ************************************************

carriageway_CD_loc = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Carriageway') & (df['loc'] == True)
                         & (df['type'] == 'JB')]

index = carriageway_CD_loc.index
num_of_rows_cw_CD_loc = carriageway_CD_loc['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E26"] = num_of_rows_cw_CD_loc
wb.save(filename)

# *********************** footway_CD loc yes  ************************************************

footway_CD_loc = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Footway') & (df['loc'] == True)
                         & (df['type'] == 'JB')]

index = footway_CD_loc.index
num_of_rows_fw_CD_loc = footway_CD_loc['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["F26"] = num_of_rows_fw_CD_loc
wb.save(filename)

# *********************** grass verge_CD loc yes  ************************************************

grassverge_CD_loc = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Grass Verge') & (df['loc'] == True)
                         & (df['type'] == 'JB')]

index = grassverge_CD_loc.index
num_of_rows_verge_CD_loc = grassverge_CD_loc['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["G26"] = num_of_rows_verge_CD_loc
wb.save(filename)

# *********************** carriageway_MH loc no and wayleave no  ************************************************

carriageway_MH = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'ManHole')]

index = carriageway_MH.index
num_of_rows_cw_MH = carriageway_MH['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B27"] = num_of_rows_cw_MH
wb.save(filename)

# *********************** footway_MH loc no and wayleave no  ************************************************

footway_MH = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'ManHole')]

index = footway_MH.index
num_of_rows_fw_MH = footway_MH['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C27"] = num_of_rows_fw_MH
wb.save(filename)

# *********************** grass verge_MH loc no and wayleave no  ************************************************

grassverge_MH = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['type'] == 'ManHole')]

index = grassverge_MH.index
num_of_rows_verge_MH = grassverge_MH['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D27"] = num_of_rows_verge_MH
wb.save(filename)

# *********************** carriageway_MH loc no and wayleave yes  ************************************************

carriageway_MH_way = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'ManHole')]

index = carriageway_MH_way.index
num_of_rows_cw_MH_way = carriageway_MH_way['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["H27"] = num_of_rows_cw_MH_way
wb.save(filename)

# *********************** footway_MH loc no and wayleave yes  ************************************************

footway_MH_way = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'ManHole')]

index = footway_MH_way.index
num_of_rows_fw_MH_way = footway_MH_way['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["I27"] = num_of_rows_fw_MH_way
wb.save(filename)

# *********************** grass verge_MH loc no and wayleave yes  ************************************************

grassverge_MH_way = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Grass Verge') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['type'] == 'ManHole')]

index = grassverge_MH_way.index
num_of_rows_verge_MH_way = grassverge_MH_way['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J27"] = num_of_rows_verge_MH_way
wb.save(filename)

# *********************** carriageway_MH loc yes  ************************************************

carriageway_MH_loc = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Carriageway') & (df['loc'] == True)
                         & (df['type'] == 'ManHole')]

index = carriageway_MH_loc.index
num_of_rows_cw_MH_loc = carriageway_MH_loc['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E27"] = num_of_rows_cw_MH_loc
wb.save(filename)

# *********************** footway_MH loc yes  ************************************************

footway_MH_loc = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Footway') & (df['loc'] == True)
                         & (df['type'] == 'ManHole')]

index = footway_MH_loc.index
num_of_rows_fw_MH_loc = footway_MH_loc['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["F27"] = num_of_rows_fw_MH_loc
wb.save(filename)

# *********************** grass verge_MH loc yes  ************************************************

grassverge_MH_loc = df.loc[(df['state'] == 'As-built') & (df['surface'] == 'Grass Verge') & (df['loc'] == True)
                         & (df['type'] == 'ManHole')]

index = grassverge_MH_loc.index
num_of_rows_verge_MH_loc = grassverge_MH_loc['core_drill'].sum()
filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["G27"] = num_of_rows_verge_MH_loc
wb.save(filename)