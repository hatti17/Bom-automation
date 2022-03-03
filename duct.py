import pandas as pd
import numpy as np
from openpyxl import load_workbook

excel_file = 'Documents/Input/duct.xlsx'
df = pd.read_excel(excel_file)

# *********************** carriageway_S+T_modular loc no and wayleave no  *********************************************

carriageway_s_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Access & Trunk') ]

carriageway_s_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_t_modular = carriageway_s_t_modular['length'].sum()

num_of_rows_cw_s_t_modular_2x96 = carriageway_s_t_modular_2x96['length'].sum()

num_of_rows_cw_s_t_modular_total = num_of_rows_cw_s_t_modular + num_of_rows_cw_s_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B38"] = num_of_rows_cw_s_t_modular_total
wb.save(filename)

# *********************** carriageway_S+T_concrete loc no and wayleave no  *********************************************

carriageway_s_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Trunk') ]

carriageway_s_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_t_concrete = carriageway_s_t_concrete['length'].sum()

num_of_rows_cw_s_t_concrete_2x96 = carriageway_s_t_concrete_2x96['length'].sum()

num_of_rows_cw_s_t_concrete_total = num_of_rows_cw_s_t_concrete + num_of_rows_cw_s_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C38"] = num_of_rows_cw_s_t_concrete_total
wb.save(filename)

# *********************** carriageway_S+T_unmade and grassverge loc no and wayleave no  *********************************************

carriageway_s_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Trunk')]

carriageway_s_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Trunk')]

carriageway_s_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Trunk') & (df['96mm'] == '2x96')]

carriageway_s_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Trunk') & (df['96mm'] == '2x96')]


num_of_rows_cw_s_t_grassverge = carriageway_s_t_grassverge['length'].sum()

num_of_rows_cw_s_t_unmade = carriageway_s_t_unmade['length'].sum()

num_of_rows_cw_s_t_grassverge_2x96 = carriageway_s_t_grassverge_2x96['length'].sum()

num_of_rows_cw_s_t_unmade_2x96 = carriageway_s_t_unmade_2x96['length'].sum()

num_of_rows_cw_s_t_unmade_grassverge_total = num_of_rows_cw_s_t_grassverge + num_of_rows_cw_s_t_unmade + num_of_rows_cw_s_t_grassverge_2x96 + num_of_rows_cw_s_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D38"] = num_of_rows_cw_s_t_unmade_grassverge_total
wb.save(filename)

# *********************** carriageway_S+T_tarmac loc no and wayleave no  *********************************************

carriageway_s_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access & Trunk') ]

carriageway_s_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_t_tarmac = carriageway_s_t_tarmac['length'].sum()

num_of_rows_cw_s_t_tarmac_2x96 = carriageway_s_t_tarmac_2x96['length'].sum()

num_of_rows_cw_s_t_tarmac_total = num_of_rows_cw_s_t_tarmac + num_of_rows_cw_s_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E38"] = num_of_rows_cw_s_t_tarmac_total
wb.save(filename)

# *********************** carriageway_D+T_modular loc no and wayleave no  *********************************************

carriageway_d_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Distribution & Trunk') ]

carriageway_d_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_d_t_modular = carriageway_d_t_modular['length'].sum()

num_of_rows_cw_d_t_modular_2x96 = carriageway_d_t_modular_2x96['length'].sum()

num_of_rows_cw_d_t_modular_total = num_of_rows_cw_d_t_modular + num_of_rows_cw_d_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B39"] = num_of_rows_cw_d_t_modular_total
wb.save(filename)

# *********************** carriageway_D+T_concrete loc no and wayleave no  *********************************************

carriageway_d_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution & Trunk') ]

carriageway_d_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_d_t_concrete = carriageway_d_t_concrete['length'].sum()

num_of_rows_cw_d_t_concrete_2x96 = carriageway_d_t_concrete_2x96['length'].sum()

num_of_rows_cw_d_t_concrete_total = num_of_rows_cw_d_t_concrete + num_of_rows_cw_d_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C39"] = num_of_rows_cw_d_t_concrete_total
wb.save(filename)

# *********************** carriageway_D+T_unmade and grassverge loc no and wayleave no  *********************************************

carriageway_d_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution & Trunk')]

carriageway_d_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution & Trunk')]

carriageway_d_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

carriageway_d_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]


num_of_rows_cw_d_t_grassverge = carriageway_d_t_grassverge['length'].sum()

num_of_rows_cw_d_t_unmade = carriageway_d_t_unmade['length'].sum()

num_of_rows_cw_d_t_grassverge_2x96 = carriageway_d_t_grassverge_2x96['length'].sum()

num_of_rows_cw_d_t_unmade_2x96 = carriageway_d_t_unmade_2x96['length'].sum()

num_of_rows_cw_d_t_unmade_grassverge_total = num_of_rows_cw_d_t_grassverge + num_of_rows_cw_d_t_unmade + num_of_rows_cw_d_t_grassverge_2x96 + num_of_rows_cw_d_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D39"] = num_of_rows_cw_d_t_unmade_grassverge_total
wb.save(filename)

# *********************** carriageway_D+T_tarmac loc no and wayleave no  *********************************************

carriageway_d_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Distribution & Trunk') ]

carriageway_d_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_d_t_tarmac = carriageway_d_t_tarmac['length'].sum()

num_of_rows_cw_d_t_tarmac_2x96 = carriageway_d_t_tarmac_2x96['length'].sum()

num_of_rows_cw_d_t_tarmac_total = num_of_rows_cw_d_t_tarmac + num_of_rows_cw_d_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E39"] = num_of_rows_cw_d_t_tarmac_total
wb.save(filename)

# *********************** carriageway_D_modular loc no and wayleave no  *********************************************

carriageway_d_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Distribution') ]

carriageway_d_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_d_modular = carriageway_d_modular['length'].sum()

num_of_rows_cw_d_modular_2x96 = carriageway_d_modular_2x96['length'].sum()

num_of_rows_cw_d_modular_total = num_of_rows_cw_d_modular + num_of_rows_cw_d_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B40"] = num_of_rows_cw_d_modular_total
wb.save(filename)

# *********************** carriageway_D_concrete loc no and wayleave no  *********************************************

carriageway_d_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution') ]

carriageway_d_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_d_concrete = carriageway_d_concrete['length'].sum()

num_of_rows_cw_d_concrete_2x96 = carriageway_d_concrete_2x96['length'].sum()

num_of_rows_cw_d_concrete_total = num_of_rows_cw_d_concrete + num_of_rows_cw_d_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C40"] = num_of_rows_cw_d_concrete_total
wb.save(filename)

# *********************** carriageway_D_unmade and grassverge loc no and wayleave no  *********************************************

carriageway_d_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution')]

carriageway_d_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution')]

carriageway_d_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

carriageway_d_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]


num_of_rows_cw_d_grassverge = carriageway_d_grassverge['length'].sum()

num_of_rows_cw_d_unmade = carriageway_d_unmade['length'].sum()

num_of_rows_cw_d_grassverge_2x96 = carriageway_d_grassverge_2x96['length'].sum()

num_of_rows_cw_d_unmade_2x96 = carriageway_d_unmade_2x96['length'].sum()

num_of_rows_cw_d_unmade_grassverge_total = num_of_rows_cw_d_grassverge + num_of_rows_cw_d_unmade + num_of_rows_cw_d_grassverge_2x96 + num_of_rows_cw_d_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D40"] = num_of_rows_cw_d_unmade_grassverge_total
wb.save(filename)

# *********************** carriageway_D_tarmac loc no and wayleave no  *********************************************

carriageway_d_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Distribution') ]

carriageway_d_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_d_tarmac = carriageway_d_tarmac['length'].sum()

num_of_rows_cw_d_tarmac_2x96 = carriageway_d_tarmac_2x96['length'].sum()

num_of_rows_cw_d_tarmac_total = num_of_rows_cw_d_tarmac + num_of_rows_cw_d_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E40"] = num_of_rows_cw_d_tarmac_total
wb.save(filename)

# *********************** carriageway_S+D+T_modular loc no and wayleave no  *********************************************

carriageway_s_d_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Acces, Distribution & Trunk') ]

carriageway_s_d_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_d_t_modular = carriageway_s_d_t_modular['length'].sum()

num_of_rows_cw_s_d_t_modular_2x96 = carriageway_s_d_t_modular_2x96['length'].sum()

num_of_rows_cw_s_d_t_modular_total = num_of_rows_cw_s_d_t_modular + num_of_rows_cw_s_d_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B41"] = num_of_rows_cw_s_d_t_modular_total
wb.save(filename)

# *********************** carriageway_S+D+T_concrete loc no and wayleave no  *********************************************

carriageway_s_d_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Acces, Distribution & Trunk') ]

carriageway_s_d_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_d_t_concrete = carriageway_s_d_t_concrete['length'].sum()

num_of_rows_cw_s_d_t_concrete_2x96 = carriageway_s_d_t_concrete_2x96['length'].sum()

num_of_rows_cw_s_d_t_concrete_total = num_of_rows_cw_s_d_t_concrete + num_of_rows_cw_s_d_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C41"] = num_of_rows_cw_s_d_t_concrete_total
wb.save(filename)

# *********************** carriageway_S+D+T_unmade and grassverge loc no and wayleave no  *********************************************

carriageway_s_d_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Acces, Distribution & Trunk')]

carriageway_s_d_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Acces, Distribution & Trunk')]

carriageway_s_d_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

carriageway_s_d_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]


num_of_rows_cw_s_d_t_grassverge = carriageway_s_d_t_grassverge['length'].sum()

num_of_rows_cw_s_d_t_unmade = carriageway_s_d_t_unmade['length'].sum()

num_of_rows_cw_s_d_t_grassverge_2x96 = carriageway_s_d_t_grassverge_2x96['length'].sum()

num_of_rows_cw_s_d_t_unmade_2x96 = carriageway_s_d_t_unmade_2x96['length'].sum()

num_of_rows_cw_s_d_t_unmade_grassverge_total = num_of_rows_cw_s_d_t_grassverge + num_of_rows_cw_s_d_t_unmade + num_of_rows_cw_s_d_t_grassverge_2x96 + num_of_rows_cw_s_d_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D41"] = num_of_rows_cw_s_d_t_unmade_grassverge_total
wb.save(filename)

# *********************** carriageway_S+D+T_tarmac loc no and wayleave no  *********************************************

carriageway_s_d_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Acces, Distribution & Trunk') ]

carriageway_s_d_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_d_t_tarmac = carriageway_s_d_t_tarmac['length'].sum()

num_of_rows_cw_s_d_t_tarmac_2x96 = carriageway_s_d_t_tarmac_2x96['length'].sum()

num_of_rows_cw_s_d_t_tarmac_total = num_of_rows_cw_s_d_t_tarmac + num_of_rows_cw_s_d_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E41"] = num_of_rows_cw_s_d_t_tarmac_total
wb.save(filename)

# *********************** carriageway_S+D_modular loc no and wayleave no  *********************************************

carriageway_s_d_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Access & Distribution') ]

carriageway_s_d_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_d_modular = carriageway_s_d_modular['length'].sum()

num_of_rows_cw_s_d_modular_2x96 = carriageway_s_d_modular_2x96['length'].sum()

num_of_rows_cw_s_d_modular_total = num_of_rows_cw_s_d_modular + num_of_rows_cw_s_d_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B42"] = num_of_rows_cw_s_d_modular_total
wb.save(filename)

# *********************** carriageway_S+D_concrete loc no and wayleave no  *********************************************

carriageway_s_d_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Distribution') ]

carriageway_s_d_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_d_concrete = carriageway_s_d_concrete['length'].sum()

num_of_rows_cw_s_d_concrete_2x96 = carriageway_s_d_concrete_2x96['length'].sum()

num_of_rows_cw_s_d_concrete_total = num_of_rows_cw_s_d_concrete + num_of_rows_cw_s_d_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C42"] = num_of_rows_cw_s_d_concrete_total
wb.save(filename)

# *********************** carriageway_S+D_unmade and grassverge loc no and wayleave no  *********************************************

carriageway_s_d_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Distribution')]

carriageway_s_d_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Distribution')]

carriageway_s_d_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

carriageway_s_d_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]


num_of_rows_cw_s_d_grassverge = carriageway_s_d_grassverge['length'].sum()

num_of_rows_cw_s_d_unmade = carriageway_s_d_unmade['length'].sum()

num_of_rows_cw_s_d_grassverge_2x96 = carriageway_s_d_grassverge_2x96['length'].sum()

num_of_rows_cw_s_d_unmade_2x96 = carriageway_s_d_unmade_2x96['length'].sum()

num_of_rows_cw_s_d_unmade_grassverge_total = num_of_rows_cw_s_d_grassverge + num_of_rows_cw_s_d_unmade + num_of_rows_cw_s_d_grassverge_2x96 + num_of_rows_cw_s_d_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D42"] = num_of_rows_cw_s_d_unmade_grassverge_total
wb.save(filename)

# *********************** carriageway_S+D_tarmac loc no and wayleave no  *********************************************

carriageway_s_d_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access & Distribution') ]

carriageway_s_d_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_d_tarmac = carriageway_s_d_tarmac['length'].sum()

num_of_rows_cw_s_d_tarmac_2x96 = carriageway_s_d_tarmac_2x96['length'].sum()

num_of_rows_cw_s_d_tarmac_total = num_of_rows_cw_s_d_tarmac + num_of_rows_cw_s_d_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E42"] = num_of_rows_cw_s_d_tarmac_total
wb.save(filename)

# *********************** carriageway_S_modular loc no and wayleave no  *********************************************

carriageway_s_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Access') ]

carriageway_s_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_modular = carriageway_s_modular['length'].sum()

num_of_rows_cw_s_modular_2x96 = carriageway_s_modular_2x96['length'].sum()

num_of_rows_cw_s_modular_total = num_of_rows_cw_s_modular + num_of_rows_cw_s_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B43"] = num_of_rows_cw_s_modular_total
wb.save(filename)

# *********************** carriageway_S_concrete loc no and wayleave no  *********************************************

carriageway_s_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access') ]

carriageway_s_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_concrete = carriageway_s_concrete['length'].sum()

num_of_rows_cw_s_concrete_2x96 = carriageway_s_concrete_2x96['length'].sum()

num_of_rows_cw_s_concrete_total = num_of_rows_cw_s_concrete + num_of_rows_cw_s_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C43"] = num_of_rows_cw_s_concrete_total
wb.save(filename)

# *********************** carriageway_S_unmade and grassverge loc no and wayleave no  *********************************************

carriageway_s_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access')]

carriageway_s_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access')]

carriageway_s_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

carriageway_s_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]


num_of_rows_cw_s_grassverge = carriageway_s_grassverge['length'].sum()

num_of_rows_cw_s_unmade = carriageway_s_unmade['length'].sum()

num_of_rows_cw_s_grassverge_2x96 = carriageway_s_grassverge_2x96['length'].sum()

num_of_rows_cw_s_unmade_2x96 = carriageway_s_unmade_2x96['length'].sum()

num_of_rows_cw_s_unmade_grassverge_total = num_of_rows_cw_s_grassverge + num_of_rows_cw_s_unmade + num_of_rows_cw_s_grassverge_2x96 + num_of_rows_cw_s_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D43"] = num_of_rows_cw_s_unmade_grassverge_total
wb.save(filename)

# *********************** carriageway_S_tarmac loc no and wayleave no  *********************************************

carriageway_s_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access') ]

carriageway_s_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_tarmac = carriageway_s_tarmac['length'].sum()

num_of_rows_cw_s_tarmac_2x96 = carriageway_s_tarmac_2x96['length'].sum()

num_of_rows_cw_s_tarmac_total = num_of_rows_cw_s_tarmac + num_of_rows_cw_s_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E43"] = num_of_rows_cw_s_tarmac_total
wb.save(filename)

# *********************** carriageway_T_modular loc no and wayleave no  *********************************************

carriageway_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Trunk') ]

carriageway_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_t_modular = carriageway_t_modular['length'].sum()

num_of_rows_cw_t_modular_2x96 = carriageway_t_modular_2x96['length'].sum()

num_of_rows_cw_t_modular_total = num_of_rows_cw_t_modular + num_of_rows_cw_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B44"] = num_of_rows_cw_t_modular_total
wb.save(filename)

# *********************** carriageway_T_concrete loc no and wayleave no  *********************************************

carriageway_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Trunk') ]

carriageway_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_t_concrete = carriageway_t_concrete['length'].sum()

num_of_rows_cw_t_concrete_2x96 = carriageway_t_concrete_2x96['length'].sum()

num_of_rows_cw_t_concrete_total = num_of_rows_cw_t_concrete + num_of_rows_cw_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C44"] = num_of_rows_cw_t_concrete_total
wb.save(filename)

# *********************** carriageway_t_unmade and grassverge loc no and wayleave no  *********************************************

carriageway_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Trunk')]

carriageway_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Trunk')]

carriageway_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

carriageway_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]


num_of_rows_cw_t_grassverge = carriageway_t_grassverge['length'].sum()

num_of_rows_cw_t_unmade = carriageway_t_unmade['length'].sum()

num_of_rows_cw_t_grassverge_2x96 = carriageway_t_grassverge_2x96['length'].sum()

num_of_rows_cw_t_unmade_2x96 = carriageway_t_unmade_2x96['length'].sum()

num_of_rows_cw_t_unmade_grassverge_total = num_of_rows_cw_t_grassverge + num_of_rows_cw_t_unmade + num_of_rows_cw_t_grassverge_2x96 + num_of_rows_cw_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D44"] = num_of_rows_cw_t_unmade_grassverge_total
wb.save(filename)

# *********************** carriageway_T_tarmac loc no and wayleave no  *********************************************

carriageway_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Trunk') ]

carriageway_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_t_tarmac = carriageway_t_tarmac['length'].sum()

num_of_rows_cw_t_tarmac_2x96 = carriageway_t_tarmac_2x96['length'].sum()

num_of_rows_cw_t_tarmac_total = num_of_rows_cw_t_tarmac + num_of_rows_cw_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E44"] = num_of_rows_cw_t_tarmac_total
wb.save(filename)

# *********************** footway_S+T_modular loc no and wayleave no  *********************************************

footway_s_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Access & Trunk') ]

footway_s_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_t_modular = footway_s_t_modular['length'].sum()

num_of_rows_fw_s_t_modular_2x96 = footway_s_t_modular_2x96['length'].sum()

num_of_rows_fw_s_t_modular_total = num_of_rows_fw_s_t_modular + num_of_rows_fw_s_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B49"] = num_of_rows_fw_s_t_modular_total
wb.save(filename)

# *********************** footway_S+T_concrete loc no and wayleave no  *********************************************

footway_s_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Trunk') ]

footway_s_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_t_concrete = footway_s_t_concrete['length'].sum()

num_of_rows_fw_s_t_concrete_2x96 = footway_s_t_concrete_2x96['length'].sum()

num_of_rows_fw_s_t_concrete_total = num_of_rows_fw_s_t_concrete + num_of_rows_fw_s_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C49"] = num_of_rows_fw_s_t_concrete_total
wb.save(filename)

# *********************** footway_S+T_unmade and grassverge loc no and wayleave no  *********************************************

footway_s_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Trunk')]

footway_s_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Trunk')]

footway_s_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Trunk') & (df['96mm'] == '2x96')]

footway_s_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Trunk') & (df['96mm'] == '2x96')]


num_of_rows_fw_s_t_grassverge = footway_s_t_grassverge['length'].sum()

num_of_rows_fw_s_t_unmade = footway_s_t_unmade['length'].sum()

num_of_rows_fw_s_t_grassverge_2x96 = footway_s_t_grassverge_2x96['length'].sum()

num_of_rows_fw_s_t_unmade_2x96 = footway_s_t_unmade_2x96['length'].sum()

num_of_rows_fw_s_t_unmade_grassverge_total = num_of_rows_fw_s_t_grassverge + num_of_rows_fw_s_t_unmade + num_of_rows_fw_s_t_grassverge_2x96 + num_of_rows_fw_s_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D49"] = num_of_rows_fw_s_t_unmade_grassverge_total
wb.save(filename)

# *********************** footway_S+T_tarmac loc no and wayleave no  *********************************************

footway_s_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access & Trunk') ]

footway_s_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_t_tarmac = footway_s_t_tarmac['length'].sum()

num_of_rows_fw_s_t_tarmac_2x96 = footway_s_t_tarmac_2x96['length'].sum()

num_of_rows_fw_s_t_tarmac_total = num_of_rows_fw_s_t_tarmac + num_of_rows_fw_s_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E49"] = num_of_rows_fw_s_t_tarmac_total
wb.save(filename)

# *********************** footway_D+T_modular loc no and wayleave no  *********************************************

footway_d_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Distribution & Trunk') ]

footway_d_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_d_t_modular = footway_d_t_modular['length'].sum()

num_of_rows_fw_d_t_modular_2x96 = footway_d_t_modular_2x96['length'].sum()

num_of_rows_fw_d_t_modular_total = num_of_rows_fw_d_t_modular + num_of_rows_fw_d_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B50"] = num_of_rows_fw_d_t_modular_total
wb.save(filename)

# *********************** footway_D+T_concrete loc no and wayleave no  *********************************************

footway_d_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution & Trunk') ]

footway_d_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_d_t_concrete = footway_d_t_concrete['length'].sum()

num_of_rows_fw_d_t_concrete_2x96 = footway_d_t_concrete_2x96['length'].sum()

num_of_rows_fw_d_t_concrete_total = num_of_rows_fw_d_t_concrete + num_of_rows_fw_d_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C50"] = num_of_rows_fw_d_t_concrete_total
wb.save(filename)

# *********************** footway_D+T_unmade and grassverge loc no and wayleave no  *********************************************

footway_d_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution & Trunk')]

footway_d_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution & Trunk')]

footway_d_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution & Trunk') & (df['96mm'] == '2x96')]

footway_d_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution & Trunk') & (df['96mm'] == '2x96')]


num_of_rows_fw_d_t_grassverge = footway_d_t_grassverge['length'].sum()

num_of_rows_fw_d_t_unmade = footway_d_t_unmade['length'].sum()

num_of_rows_fw_d_t_grassverge_2x96 = footway_d_t_grassverge_2x96['length'].sum()

num_of_rows_fw_d_t_unmade_2x96 = footway_d_t_unmade_2x96['length'].sum()

num_of_rows_fw_d_t_unmade_grassverge_total = num_of_rows_fw_d_t_grassverge + num_of_rows_fw_d_t_unmade + num_of_rows_fw_d_t_grassverge_2x96 + num_of_rows_fw_d_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D50"] = num_of_rows_fw_d_t_unmade_grassverge_total
wb.save(filename)

# *********************** footway_D+T_tarmac loc no and wayleave no  *********************************************

footway_d_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Distribution & Trunk') ]

footway_d_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_d_t_tarmac = footway_d_t_tarmac['length'].sum()

num_of_rows_fw_d_t_tarmac_2x96 = footway_d_t_tarmac_2x96['length'].sum()

num_of_rows_fw_d_t_tarmac_total = num_of_rows_fw_d_t_tarmac + num_of_rows_fw_d_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E50"] = num_of_rows_fw_d_t_tarmac_total
wb.save(filename)

# *********************** footway_D_modular loc no and wayleave no  *********************************************

footway_d_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Distribution') ]

footway_d_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_d_modular = footway_d_modular['length'].sum()

num_of_rows_fw_d_modular_2x96 = footway_d_modular_2x96['length'].sum()

num_of_rows_fw_d_modular_total = num_of_rows_fw_d_modular + num_of_rows_fw_d_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B51"] = num_of_rows_fw_d_modular_total
wb.save(filename)


# *********************** footway_D_concrete loc no and wayleave no  *********************************************

footway_d_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution') ]

footway_d_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_d_concrete = footway_d_concrete['length'].sum()

num_of_rows_fw_d_concrete_2x96 = footway_d_concrete_2x96['length'].sum()

num_of_rows_fw_d_concrete_total = num_of_rows_fw_d_concrete + num_of_rows_fw_d_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C51"] = num_of_rows_fw_d_concrete_total
wb.save(filename)

# *********************** footway_D_unmade and grassverge loc no and wayleave no  *********************************************

footway_d_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution')]

footway_d_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution')]

footway_d_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution') & (df['96mm'] == '2x96')]

footway_d_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution') & (df['96mm'] == '2x96')]


num_of_rows_fw_d_grassverge = footway_d_grassverge['length'].sum()

num_of_rows_fw_d_unmade = footway_d_unmade['length'].sum()

num_of_rows_fw_d_grassverge_2x96 = footway_d_grassverge_2x96['length'].sum()

num_of_rows_fw_d_unmade_2x96 = footway_d_unmade_2x96['length'].sum()

num_of_rows_fw_d_unmade_grassverge_total = num_of_rows_fw_d_grassverge + num_of_rows_fw_d_unmade + num_of_rows_fw_d_grassverge_2x96 + num_of_rows_fw_d_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D51"] = num_of_rows_fw_d_unmade_grassverge_total
wb.save(filename)

# *********************** footway_D_tarmac loc no and wayleave no  *********************************************

footway_d_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Distribution') ]

footway_d_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_d_tarmac = footway_d_tarmac['length'].sum()

num_of_rows_fw_d_tarmac_2x96 = footway_d_tarmac_2x96['length'].sum()

num_of_rows_fw_d_tarmac_total = num_of_rows_fw_d_tarmac + num_of_rows_fw_d_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E51"] = num_of_rows_fw_d_tarmac_total
wb.save(filename)

# *********************** footway_S_D_T_modular loc no and wayleave no  *********************************************

footway_s_d_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Acces, Distribution & Trunk') ]

footway_s_d_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_d_t_modular = footway_s_d_t_modular['length'].sum()

num_of_rows_fw_s_d_t_modular_2x96 = footway_s_d_t_modular_2x96['length'].sum()

num_of_rows_fw_s_d_t_modular_total = num_of_rows_fw_s_d_t_modular + num_of_rows_fw_s_d_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B52"] = num_of_rows_fw_s_d_t_modular_total
wb.save(filename)


# *********************** footway_S+D+T_concrete loc no and wayleave no  *********************************************

footway_s_d_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Acces, Distribution & Trunk') ]

footway_s_d_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_d_t_concrete = footway_s_d_t_concrete['length'].sum()

num_of_rows_fw_s_d_t_concrete_2x96 = footway_s_d_t_concrete_2x96['length'].sum()

num_of_rows_fw_s_d_t_concrete_total = num_of_rows_fw_s_d_t_concrete + num_of_rows_fw_s_d_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C52"] = num_of_rows_fw_s_d_t_concrete_total
wb.save(filename)

# *********************** footway_S+D+T_unmade and grassverge loc no and wayleave no  *********************************************

footway_s_d_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Acces, Distribution & Trunk')]

footway_s_d_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Acces, Distribution & Trunk')]

footway_s_d_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Acces, Distribution & Trunk') & (df['96mm'] == '2x96')]

footway_s_d_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Acces, Distribution & Trunk') & (df['96mm'] == '2x96')]


num_of_rows_fw_s_d_t_grassverge = footway_s_d_t_grassverge['length'].sum()

num_of_rows_fw_s_d_t_unmade = footway_s_d_t_unmade['length'].sum()

num_of_rows_fw_s_d_t_grassverge_2x96 = footway_s_d_t_grassverge_2x96['length'].sum()

num_of_rows_fw_s_d_t_unmade_2x96 = footway_s_d_t_unmade_2x96['length'].sum()

num_of_rows_fw_s_d_t_unmade_grassverge_total = num_of_rows_fw_s_d_t_grassverge + num_of_rows_fw_s_d_t_unmade + num_of_rows_fw_s_d_t_grassverge_2x96 + num_of_rows_fw_s_d_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D52"] = num_of_rows_fw_s_d_t_unmade_grassverge_total
wb.save(filename)

# *********************** footway_S+D+T_tarmac loc no and wayleave no  *********************************************

footway_s_d_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Acces, Distribution & Trunk') ]

footway_s_d_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_d_t_tarmac = footway_s_d_t_tarmac['length'].sum()

num_of_rows_fw_s_d_t_tarmac_2x96 = footway_s_d_t_tarmac_2x96['length'].sum()

num_of_rows_fw_s_d_t_tarmac_total = num_of_rows_fw_s_d_t_tarmac + num_of_rows_fw_s_d_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E52"] = num_of_rows_fw_s_d_t_tarmac_total
wb.save(filename)

# *********************** footway_S+D_modular loc no and wayleave no  *********************************************

footway_s_d_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Access & Distribution') ]

footway_s_d_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_d_modular = footway_s_d_modular['length'].sum()

num_of_rows_fw_s_d_modular_2x96 = footway_s_d_modular_2x96['length'].sum()

num_of_rows_fw_s_d_modular_total = num_of_rows_fw_s_d_modular + num_of_rows_fw_s_d_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B53"] = num_of_rows_fw_s_d_modular_total
wb.save(filename)


# *********************** footway_S+D_concrete loc no and wayleave no  *********************************************

footway_s_d_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Distribution') ]

footway_s_d_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_d_concrete = footway_s_d_concrete['length'].sum()

num_of_rows_fw_s_d_concrete_2x96 = footway_s_d_concrete_2x96['length'].sum()

num_of_rows_fw_s_d_concrete_total = num_of_rows_fw_s_d_concrete + num_of_rows_fw_s_d_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C53"] = num_of_rows_fw_s_d_concrete_total
wb.save(filename)

# *********************** footway_S+D_unmade and grassverge loc no and wayleave no  *********************************************

footway_s_d_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Distribution')]

footway_s_d_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Distribution')]

footway_s_d_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Distribution') & (df['96mm'] == '2x96')]

footway_s_d_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Distribution') & (df['96mm'] == '2x96')]


num_of_rows_fw_s_d_grassverge = footway_s_d_grassverge['length'].sum()

num_of_rows_fw_s_d_unmade = footway_s_d_unmade['length'].sum()

num_of_rows_fw_s_d_grassverge_2x96 = footway_s_d_grassverge_2x96['length'].sum()

num_of_rows_fw_s_d_unmade_2x96 = footway_s_d_unmade_2x96['length'].sum()

num_of_rows_fw_s_d_unmade_grassverge_total = num_of_rows_fw_s_d_grassverge + num_of_rows_fw_s_d_unmade + num_of_rows_fw_s_d_grassverge_2x96 + num_of_rows_fw_s_d_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D53"] = num_of_rows_fw_s_d_unmade_grassverge_total
wb.save(filename)

# *********************** footway_S+D_tarmac loc no and wayleave no  *********************************************

footway_s_d_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access & Distribution') ]

footway_s_d_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_d_tarmac = footway_s_d_tarmac['length'].sum()

num_of_rows_fw_s_d_tarmac_2x96 = footway_s_d_tarmac_2x96['length'].sum()

num_of_rows_fw_s_d_tarmac_total = num_of_rows_fw_s_d_tarmac + num_of_rows_fw_s_d_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E53"] = num_of_rows_fw_s_d_tarmac_total
wb.save(filename)

# *********************** footway_S_modular loc no and wayleave no  *********************************************

footway_s_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Access') ]

footway_s_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_modular = footway_s_modular['length'].sum()

num_of_rows_fw_s_modular_2x96 = footway_s_modular_2x96['length'].sum()

num_of_rows_fw_s_modular_total = num_of_rows_fw_s_modular + num_of_rows_fw_s_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B54"] = num_of_rows_fw_s_modular_total
wb.save(filename)


# *********************** footway_S_concrete loc no and wayleave no  *********************************************

footway_s_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access') ]

footway_s_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_concrete = footway_s_concrete['length'].sum()

num_of_rows_fw_s_concrete_2x96 = footway_s_concrete_2x96['length'].sum()

num_of_rows_fw_s_concrete_total = num_of_rows_fw_s_concrete + num_of_rows_fw_s_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C54"] = num_of_rows_fw_s_concrete_total
wb.save(filename)

# *********************** footway_S_unmade and grassverge loc no and wayleave no  *********************************************

footway_s_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access')]

footway_s_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access')]

footway_s_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access') & (df['96mm'] == '2x96')]

footway_s_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access') & (df['96mm'] == '2x96')]


num_of_rows_fw_s_grassverge = footway_s_grassverge['length'].sum()

num_of_rows_fw_s_unmade = footway_s_unmade['length'].sum()

num_of_rows_fw_s_grassverge_2x96 = footway_s_grassverge_2x96['length'].sum()

num_of_rows_fw_s_unmade_2x96 = footway_s_unmade_2x96['length'].sum()

num_of_rows_fw_s_unmade_grassverge_total = num_of_rows_fw_s_grassverge + num_of_rows_fw_s_unmade + num_of_rows_fw_s_grassverge_2x96 + num_of_rows_fw_s_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D54"] = num_of_rows_fw_s_unmade_grassverge_total
wb.save(filename)

# *********************** footway_S_tarmac loc no and wayleave no  ************************************************

footway_s_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Access') ]

footway_s_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Access')
                         &  (df['96mm'] == '2x96') ]

num_of_rows_fw_s_tarmac = footway_s_tarmac['length'].sum()

num_of_rows_fw_s_tarmac_2x96 = footway_s_tarmac_2x96['length'].sum()

num_of_rows_fw_s_tarmac_total = num_of_rows_fw_s_tarmac + num_of_rows_fw_s_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E54"] = num_of_rows_fw_s_tarmac_total
wb.save(filename)

# *********************** footway_T_modular loc no and wayleave no  *********************************************

footway_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Trunk') ]

footway_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_t_modular = footway_t_modular['length'].sum()

num_of_rows_fw_t_modular_2x96 = footway_t_modular_2x96['length'].sum()

num_of_rows_fw_t_modular_total = num_of_rows_fw_t_modular + num_of_rows_fw_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B55"] = num_of_rows_fw_t_modular_total
wb.save(filename)


# *********************** footway_t_concrete loc no and wayleave no  *********************************************

footway_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Trunk') ]

footway_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_t_concrete = footway_t_concrete['length'].sum()

num_of_rows_fw_t_concrete_2x96 = footway_t_concrete_2x96['length'].sum()

num_of_rows_fw_t_concrete_total = num_of_rows_fw_t_concrete + num_of_rows_fw_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C55"] = num_of_rows_fw_t_concrete_total
wb.save(filename)

# *********************** footway_T_unmade and grassverge loc no and wayleave no  *********************************************

footway_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Trunk')]

footway_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Trunk')]

footway_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Trunk') & (df['96mm'] == '2x96')]

footway_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Trunk') & (df['96mm'] == '2x96')]


num_of_rows_fw_t_grassverge = footway_t_grassverge['length'].sum()

num_of_rows_fw_t_unmade = footway_t_unmade['length'].sum()

num_of_rows_fw_t_grassverge_2x96 = footway_t_grassverge_2x96['length'].sum()

num_of_rows_fw_t_unmade_2x96 = footway_t_unmade_2x96['length'].sum()

num_of_rows_fw_t_unmade_grassverge_total = num_of_rows_fw_t_grassverge + num_of_rows_fw_t_unmade + num_of_rows_fw_t_grassverge_2x96 + num_of_rows_fw_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D55"] = num_of_rows_fw_t_unmade_grassverge_total
wb.save(filename)

# *********************** footway_T_tarmac loc no and wayleave no  ************************************************

footway_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Trunk') ]

footway_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Trunk')
                         &  (df['96mm'] == '2x96') ]

num_of_rows_fw_t_tarmac = footway_t_tarmac['length'].sum()

num_of_rows_fw_t_tarmac_2x96 = footway_t_tarmac_2x96['length'].sum()

num_of_rows_fw_t_tarmac_total = num_of_rows_fw_t_tarmac + num_of_rows_fw_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E55"] = num_of_rows_fw_t_tarmac_total
wb.save(filename)

# *********************** road crossing_S+T_modular loc no and wayleave no  *********************************************

roadcrossing_s_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Access & Trunk') ]

roadcrossing_s_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_s_t_modular = roadcrossing_s_t_modular['length'].sum()

num_of_rows_rc_s_t_modular_2x96 = roadcrossing_s_t_modular_2x96['length'].sum()

num_of_rows_rc_s_t_modular_total = num_of_rows_rc_s_t_modular + num_of_rows_rc_s_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B60"] = num_of_rows_rc_s_t_modular_total
wb.save(filename)

# *********************** road crossing_S+T_concrete loc no and wayleave no  *********************************************

roadcrossing_s_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Trunk') ]

roadcrossing_s_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_s_t_concrete = roadcrossing_s_t_concrete['length'].sum()

num_of_rows_rc_s_t_concrete_2x96 = roadcrossing_s_t_concrete_2x96['length'].sum()

num_of_rows_rc_s_t_concrete_total = num_of_rows_rc_s_t_concrete + num_of_rows_rc_s_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C60"] = num_of_rows_rc_s_t_concrete_total
wb.save(filename)

# *********************** road crossing_S+T_unmade and grassverge loc no and wayleave no  *********************************************

roadcrossing_s_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Trunk')]

roadcrossing_s_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Trunk')]

roadcrossing_s_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Trunk') & (df['96mm'] == '2x96')]

roadcrossing_s_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Trunk') & (df['96mm'] == '2x96')]


num_of_rows_rc_s_t_grassverge = roadcrossing_s_t_grassverge['length'].sum()

num_of_rows_rc_s_t_unmade = roadcrossing_s_t_unmade['length'].sum()

num_of_rows_rc_s_t_grassverge_2x96 = roadcrossing_s_t_grassverge_2x96['length'].sum()

num_of_rows_rc_s_t_unmade_2x96 = roadcrossing_s_t_unmade_2x96['length'].sum()

num_of_rows_rc_s_t_unmade_grassverge_total = num_of_rows_rc_s_t_grassverge + num_of_rows_rc_s_t_unmade + num_of_rows_rc_s_t_grassverge_2x96 + num_of_rows_rc_s_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D60"] = num_of_rows_rc_s_t_unmade_grassverge_total
wb.save(filename)

# *********************** road crossing_S+T_tarmac loc no and wayleave no  *********************************************

roadcrossing_s_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access & Trunk') ]

roadcrossing_s_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_s_t_tarmac = roadcrossing_s_t_tarmac['length'].sum()

num_of_rows_rc_s_t_tarmac_2x96 = roadcrossing_s_t_tarmac_2x96['length'].sum()

num_of_rows_rc_s_t_tarmac_total = num_of_rows_rc_s_t_tarmac + num_of_rows_rc_s_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E60"] = num_of_rows_rc_s_t_tarmac_total
wb.save(filename)

# *********************** road crossing_D+T_modular loc no and wayleave no  *********************************************

roadcrossing_d_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Distribution & Trunk') ]

roadcrossing_d_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_d_t_modular = roadcrossing_d_t_modular['length'].sum()

num_of_rows_rc_d_t_modular_2x96 = roadcrossing_d_t_modular_2x96['length'].sum()

num_of_rows_rc_d_t_modular_total = num_of_rows_rc_d_t_modular + num_of_rows_rc_d_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B61"] = num_of_rows_rc_d_t_modular_total
wb.save(filename)

# *********************** road crossing_D+T_concrete loc no and wayleave no  *********************************************

roadcrossing_d_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution & Trunk') ]

roadcrossing_d_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_d_t_concrete = roadcrossing_d_t_concrete['length'].sum()

num_of_rows_rc_d_t_concrete_2x96 = roadcrossing_d_t_concrete_2x96['length'].sum()

num_of_rows_rc_d_t_concrete_total = num_of_rows_rc_d_t_concrete + num_of_rows_rc_d_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C61"] = num_of_rows_rc_d_t_concrete_total
wb.save(filename)

# *********************** road crossing_D+T_unmade and grassverge loc no and wayleave no  *********************************************

roadcrossing_d_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution & Trunk')]

roadcrossing_d_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution & Trunk')]

roadcrossing_d_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution & Trunk') & (df['96mm'] == '2x96')]

roadcrossing_d_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution & Trunk') & (df['96mm'] == '2x96')]


num_of_rows_rc_d_t_grassverge = roadcrossing_d_t_grassverge['length'].sum()

num_of_rows_rc_d_t_unmade = roadcrossing_d_t_unmade['length'].sum()

num_of_rows_rc_d_t_grassverge_2x96 = roadcrossing_d_t_grassverge_2x96['length'].sum()

num_of_rows_rc_d_t_unmade_2x96 = roadcrossing_d_t_unmade_2x96['length'].sum()

num_of_rows_rc_d_t_unmade_grassverge_total = num_of_rows_rc_d_t_grassverge + num_of_rows_rc_d_t_unmade + num_of_rows_rc_d_t_grassverge_2x96 + num_of_rows_rc_d_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D61"] = num_of_rows_rc_d_t_unmade_grassverge_total
wb.save(filename)

# *********************** road crossing_D+T_tarmac loc no and wayleave no  *********************************************

roadcrossing_d_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Distribution & Trunk') ]

roadcrossing_d_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_d_t_tarmac = roadcrossing_d_t_tarmac['length'].sum()

num_of_rows_rc_d_t_tarmac_2x96 = roadcrossing_d_t_tarmac_2x96['length'].sum()

num_of_rows_rc_d_t_tarmac_total = num_of_rows_rc_d_t_tarmac + num_of_rows_rc_d_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E61"] = num_of_rows_rc_d_t_tarmac_total
wb.save(filename)

# *********************** road crossing_D_modular loc no and wayleave no  *********************************************

roadcrossing_d_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Distribution') ]

roadcrossing_d_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_d_modular = roadcrossing_d_modular['length'].sum()

num_of_rows_rc_d_modular_2x96 = roadcrossing_d_modular_2x96['length'].sum()

num_of_rows_rc_d_modular_total = num_of_rows_rc_d_modular + num_of_rows_rc_d_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B62"] = num_of_rows_rc_d_modular_total
wb.save(filename)

# *********************** road crossing_D_concrete loc no and wayleave no  *********************************************

roadcrossing_d_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution') ]

roadcrossing_d_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_d_concrete = roadcrossing_d_concrete['length'].sum()

num_of_rows_rc_d_concrete_2x96 = roadcrossing_d_concrete_2x96['length'].sum()

num_of_rows_rc_d_concrete_total = num_of_rows_rc_d_concrete + num_of_rows_rc_d_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C62"] = num_of_rows_rc_d_concrete_total
wb.save(filename)

# *********************** road crossing_D_unmade and grassverge loc no and wayleave no  *********************************************

roadcrossing_d_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution')]

roadcrossing_d_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution')]

roadcrossing_d_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution') & (df['96mm'] == '2x96')]

roadcrossing_d_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution') & (df['96mm'] == '2x96')]


num_of_rows_rc_d_grassverge = roadcrossing_d_grassverge['length'].sum()

num_of_rows_rc_d_unmade = roadcrossing_d_unmade['length'].sum()

num_of_rows_rc_d_grassverge_2x96 = roadcrossing_d_grassverge_2x96['length'].sum()

num_of_rows_rc_d_unmade_2x96 = roadcrossing_d_unmade_2x96['length'].sum()

num_of_rows_rc_d_unmade_grassverge_total = num_of_rows_rc_d_grassverge + num_of_rows_rc_d_t_unmade + num_of_rows_rc_d_grassverge_2x96 + num_of_rows_rc_d_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D62"] = num_of_rows_rc_d_unmade_grassverge_total
wb.save(filename)

# *********************** road crossing_D_tarmac loc no and wayleave no  *********************************************

roadcrossing_d_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Distribution') ]

roadcrossing_d_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_d_tarmac = roadcrossing_d_tarmac['length'].sum()

num_of_rows_rc_d_tarmac_2x96 = roadcrossing_d_tarmac_2x96['length'].sum()

num_of_rows_rc_d_tarmac_total = num_of_rows_rc_d_tarmac + num_of_rows_rc_d_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E62"] = num_of_rows_rc_d_tarmac_total
wb.save(filename)

# *********************** road crossing_S_D_T_modular loc no and wayleave no  *********************************************

roadcrossing_s_d_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Acces, Distribution & Trunk') ]

roadcrossing_s_d_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_s_d_t_modular = roadcrossing_s_d_t_modular['length'].sum()

num_of_rows_rc_s_d_t_modular_2x96 = roadcrossing_s_d_t_modular_2x96['length'].sum()

num_of_rows_rc_s_d_t_modular_total = num_of_rows_rc_s_d_t_modular + num_of_rows_rc_s_d_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B63"] = num_of_rows_rc_s_d_t_modular_total
wb.save(filename)

# *********************** road crossing_S_D_T_concrete loc no and wayleave no  *********************************************

roadcrossing_s_d_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Acces, Distribution & Trunk') ]

roadcrossing_s_d_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_s_d_t_concrete = roadcrossing_s_d_t_concrete['length'].sum()

num_of_rows_rc_s_d_t_concrete_2x96 = roadcrossing_s_d_t_concrete_2x96['length'].sum()

num_of_rows_rc_s_d_t_concrete_total = num_of_rows_rc_s_d_t_concrete + num_of_rows_rc_s_d_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C63"] = num_of_rows_rc_s_d_t_concrete_total
wb.save(filename)

# *********************** road crossing_S_D_T_unmade and grassverge loc no and wayleave no  *********************************************

roadcrossing_s_d_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Acces, Distribution & Trunk')]

roadcrossing_s_d_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Acces, Distribution & Trunk')]

roadcrossing_s_d_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Acces, Distribution & Trunk') & (df['96mm'] == '2x96')]

roadcrossing_s_d_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Acces, Distribution & Trunk') & (df['96mm'] == '2x96')]


num_of_rows_rc_s_d_t_grassverge = roadcrossing_s_d_t_grassverge['length'].sum()

num_of_rows_rc_s_d_t_unmade = roadcrossing_s_d_t_unmade['length'].sum()

num_of_rows_rc_s_d_t_grassverge_2x96 = roadcrossing_s_d_t_grassverge_2x96['length'].sum()

num_of_rows_rc_s_d_t_unmade_2x96 = roadcrossing_s_d_t_unmade_2x96['length'].sum()

num_of_rows_rc_s_d_t_unmade_grassverge_total = num_of_rows_rc_s_d_t_grassverge + num_of_rows_rc_s_d_t_unmade + num_of_rows_rc_s_d_t_grassverge_2x96 + num_of_rows_rc_s_d_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D63"] = num_of_rows_rc_s_d_t_unmade_grassverge_total
wb.save(filename)

# *********************** road crossing_S_D_T_tarmac loc no and wayleave no  *********************************************

roadcrossing_s_d_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Acces, Distribution & Trunk') ]

roadcrossing_s_d_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_s_d_t_tarmac = roadcrossing_s_d_t_tarmac['length'].sum()

num_of_rows_rc_s_d_t_tarmac_2x96 = roadcrossing_s_d_t_tarmac_2x96['length'].sum()

num_of_rows_rc_s_d_t_tarmac_total = num_of_rows_rc_s_d_t_tarmac + num_of_rows_rc_s_d_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E63"] = num_of_rows_rc_s_d_t_tarmac_total
wb.save(filename)

# *********************** road crossing_S_D_modular loc no and wayleave no  *********************************************

roadcrossing_s_d_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Access & Distribution') ]

roadcrossing_s_d_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_s_d_modular = roadcrossing_s_d_modular['length'].sum()

num_of_rows_rc_s_d_modular_2x96 = roadcrossing_s_d_modular_2x96['length'].sum()

num_of_rows_rc_s_d_modular_total = num_of_rows_rc_s_d_modular + num_of_rows_rc_s_d_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B64"] = num_of_rows_rc_s_d_modular_total
wb.save(filename)

# *********************** road crossing_S_D_concrete loc no and wayleave no  *********************************************

roadcrossing_s_d_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Distribution') ]

roadcrossing_s_d_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_s_d_concrete = roadcrossing_s_d_concrete['length'].sum()

num_of_rows_rc_s_d_concrete_2x96 = roadcrossing_s_d_concrete_2x96['length'].sum()

num_of_rows_rc_s_d_concrete_total = num_of_rows_rc_s_d_concrete + num_of_rows_rc_s_d_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C64"] = num_of_rows_rc_s_d_concrete_total
wb.save(filename)

# *********************** road crossing_S_D_unmade and grassverge loc no and wayleave no  *********************************************

roadcrossing_s_d_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Distribution')]

roadcrossing_s_d_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Distribution')]

roadcrossing_s_d_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Distribution') & (df['96mm'] == '2x96')]

roadcrossing_s_d_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Distribution') & (df['96mm'] == '2x96')]


num_of_rows_rc_s_d_grassverge = roadcrossing_s_d_grassverge['length'].sum()

num_of_rows_rc_s_d_unmade = roadcrossing_s_d_unmade['length'].sum()

num_of_rows_rc_s_d_grassverge_2x96 = roadcrossing_s_d_grassverge_2x96['length'].sum()

num_of_rows_rc_s_d_unmade_2x96 = roadcrossing_s_d_unmade_2x96['length'].sum()

num_of_rows_rc_s_d_unmade_grassverge_total = num_of_rows_rc_s_d_grassverge + num_of_rows_rc_s_d_unmade + num_of_rows_rc_s_d_grassverge_2x96 + num_of_rows_rc_s_d_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D64"] = num_of_rows_rc_s_d_unmade_grassverge_total
wb.save(filename)

# *********************** road crossing_S_D_tarmac loc no and wayleave no  *********************************************

roadcrossing_s_d_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access & Distribution') ]

roadcrossing_s_d_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_s_d_tarmac = roadcrossing_s_d_tarmac['length'].sum()

num_of_rows_rc_s_d_tarmac_2x96 = roadcrossing_s_d_tarmac_2x96['length'].sum()

num_of_rows_rc_s_d_tarmac_total = num_of_rows_rc_s_d_tarmac + num_of_rows_rc_s_d_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E64"] = num_of_rows_rc_s_d_tarmac_total
wb.save(filename)

# *********************** road crossing_S_modular loc no and wayleave no  *********************************************

roadcrossing_s_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Access') ]

roadcrossing_s_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_s_modular = roadcrossing_s_modular['length'].sum()

num_of_rows_rc_s_modular_2x96 = roadcrossing_s_modular_2x96['length'].sum()

num_of_rows_rc_s_modular_total = num_of_rows_rc_s_modular + num_of_rows_rc_s_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B65"] = num_of_rows_rc_s_modular_total
wb.save(filename)

# *********************** road crossing_S_concrete loc no and wayleave no  *********************************************

roadcrossing_s_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access') ]

roadcrossing_s_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_s_concrete = roadcrossing_s_concrete['length'].sum()

num_of_rows_rc_s_concrete_2x96 = roadcrossing_s_concrete_2x96['length'].sum()

num_of_rows_rc_s_concrete_total = num_of_rows_rc_s_concrete + num_of_rows_rc_s_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C65"] = num_of_rows_rc_s_concrete_total
wb.save(filename)

# *********************** road crossing_S_unmade and grassverge loc no and wayleave no  *********************************************

roadcrossing_s_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access')]

roadcrossing_s_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access')]

roadcrossing_s_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access') & (df['96mm'] == '2x96')]

roadcrossing_s_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access') & (df['96mm'] == '2x96')]


num_of_rows_rc_s_grassverge = roadcrossing_s_grassverge['length'].sum()

num_of_rows_rc_s_unmade = roadcrossing_s_unmade['length'].sum()

num_of_rows_rc_s_grassverge_2x96 = roadcrossing_s_grassverge_2x96['length'].sum()

num_of_rows_rc_s_unmade_2x96 = roadcrossing_s_unmade_2x96['length'].sum()

num_of_rows_rc_s_unmade_grassverge_total = num_of_rows_rc_s_grassverge + num_of_rows_rc_s_unmade + num_of_rows_rc_s_grassverge_2x96 + num_of_rows_rc_s_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D65"] = num_of_rows_rc_s_unmade_grassverge_total
wb.save(filename)

# *********************** road crossing_S_tarmac loc no and wayleave no  *********************************************

roadcrossing_s_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access') ]

roadcrossing_s_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_s_tarmac = roadcrossing_s_tarmac['length'].sum()

num_of_rows_rc_s_tarmac_2x96 = roadcrossing_s_tarmac_2x96['length'].sum()

num_of_rows_rc_s_tarmac_total = num_of_rows_rc_s_tarmac + num_of_rows_rc_s_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E65"] = num_of_rows_rc_s_tarmac_total
wb.save(filename)

# *********************** road crossing_T_modular loc no and wayleave no  *********************************************

roadcrossing_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Trunk') ]

roadcrossing_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_t_modular = roadcrossing_t_modular['length'].sum()

num_of_rows_rc_t_modular_2x96 = roadcrossing_t_modular_2x96['length'].sum()

num_of_rows_rc_t_modular_total = num_of_rows_rc_t_modular + num_of_rows_rc_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B66"] = num_of_rows_rc_t_modular_total
wb.save(filename)

# *********************** road crossing_T_concrete loc no and wayleave no  *********************************************

roadcrossing_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Trunk') ]

roadcrossing_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_t_concrete = roadcrossing_t_concrete['length'].sum()

num_of_rows_rc_t_concrete_2x96 = roadcrossing_t_concrete_2x96['length'].sum()

num_of_rows_rc_t_concrete_total = num_of_rows_rc_t_concrete + num_of_rows_rc_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C66"] = num_of_rows_rc_t_concrete_total
wb.save(filename)

# *********************** road crossing_T_unmade and grassverge loc no and wayleave no  *********************************************

roadcrossing_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Trunk')]

roadcrossing_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Trunk')]

roadcrossing_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Trunk') & (df['96mm'] == '2x96')]

roadcrossing_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Trunk') & (df['96mm'] == '2x96')]


num_of_rows_rc_t_grassverge = roadcrossing_t_grassverge['length'].sum()

num_of_rows_rc_t_unmade = roadcrossing_t_unmade['length'].sum()

num_of_rows_rc_t_grassverge_2x96 = roadcrossing_t_grassverge_2x96['length'].sum()

num_of_rows_rc_t_unmade_2x96 = roadcrossing_t_unmade_2x96['length'].sum()

num_of_rows_rc_t_unmade_grassverge_total = num_of_rows_rc_t_grassverge + num_of_rows_rc_t_unmade + num_of_rows_rc_t_grassverge_2x96 + num_of_rows_rc_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D66"] = num_of_rows_rc_t_unmade_grassverge_total
wb.save(filename)

# *********************** road crossing_T_tarmac loc no and wayleave no  *********************************************

roadcrossing_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Trunk') ]

roadcrossing_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Road Crossing') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_rc_t_tarmac = roadcrossing_t_tarmac['length'].sum()

num_of_rows_rc_t_tarmac_2x96 = roadcrossing_t_tarmac_2x96['length'].sum()

num_of_rows_rc_t_tarmac_total = num_of_rows_rc_t_tarmac + num_of_rows_rc_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E66"] = num_of_rows_rc_t_tarmac_total
wb.save(filename)

# *********************** carriageway_S+T_modular loc no and wayleave yes  *********************************************

carriageway_s_t_modular_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular' ) & (df['type'] == 'Access & Trunk') ]

carriageway_s_t_modular_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_t_modular_way = carriageway_s_t_modular_way['length'].sum()

num_of_rows_cw_s_t_modular_2x96_way = carriageway_s_t_modular_2x96_way['length'].sum()

num_of_rows_cw_s_t_modular_total_way = num_of_rows_cw_s_t_modular_way + num_of_rows_cw_s_t_modular_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J38"] = num_of_rows_cw_s_t_modular_total_way
wb.save(filename)

# *********************** carriageway_S+T_concrete loc no and wayleave yes  *********************************************

carriageway_s_t_concrete_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Trunk') ]

carriageway_s_t_concrete_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_t_concrete_way = carriageway_s_t_concrete_way['length'].sum()

num_of_rows_cw_s_t_concrete_2x96_way = carriageway_s_t_concrete_2x96_way['length'].sum()

num_of_rows_cw_s_t_concrete_total_way = num_of_rows_cw_s_t_concrete_way + num_of_rows_cw_s_t_concrete_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K38"] = num_of_rows_cw_s_t_concrete_total_way
wb.save(filename)

# *********************** carriageway_S+T_unmade and grassverge loc no and wayleave yes  *********************************************

carriageway_s_t_grassverge_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Trunk')]

carriageway_s_t_unmade_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Trunk')]

carriageway_s_t_grassverge_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Trunk') & (df['96mm'] == '2x96')]

carriageway_s_t_unmade_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Trunk') & (df['96mm'] == '2x96')]


num_of_rows_cw_s_t_grassverge_way = carriageway_s_t_grassverge_way['length'].sum()

num_of_rows_cw_s_t_unmade_way = carriageway_s_t_unmade_way['length'].sum()

num_of_rows_cw_s_t_grassverge_2x96_way = carriageway_s_t_grassverge_2x96_way['length'].sum()

num_of_rows_cw_s_t_unmade_2x96_way = carriageway_s_t_unmade_2x96_way['length'].sum()

num_of_rows_cw_s_t_unmade_grassverge_total_way = num_of_rows_cw_s_t_grassverge_way + num_of_rows_cw_s_t_unmade_way + num_of_rows_cw_s_t_grassverge_2x96_way + num_of_rows_cw_s_t_unmade_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L38"] = num_of_rows_cw_s_t_unmade_grassverge_total_way
wb.save(filename)

# *********************** carriageway_S+T_tarmac loc no and wayleave yes  *********************************************

carriageway_s_t_tarmac_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access & Trunk') ]

carriageway_s_t_tarmac_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_t_tarmac_way = carriageway_s_t_tarmac_way['length'].sum()

num_of_rows_cw_s_t_tarmac_2x96_way = carriageway_s_t_tarmac_2x96_way['length'].sum()

num_of_rows_cw_s_t_tarmac_total_way = num_of_rows_cw_s_t_tarmac_way + num_of_rows_cw_s_t_tarmac_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M38"] = num_of_rows_cw_s_t_tarmac_total_way
wb.save(filename)

# *********************** carriageway_D+T_modular loc no and wayleave yes  *********************************************

carriageway_d_t_modular_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular' ) & (df['type'] == 'Distribution & Trunk') ]

carriageway_d_t_modular_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_d_t_modular_way = carriageway_d_t_modular_way['length'].sum()

num_of_rows_cw_d_t_modular_2x96_way = carriageway_d_t_modular_2x96_way['length'].sum()

num_of_rows_cw_d_t_modular_total_way = num_of_rows_cw_d_t_modular_way + num_of_rows_cw_d_t_modular_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J39"] = num_of_rows_cw_d_t_modular_total_way
wb.save(filename)

# *********************** carriageway_D+T_concrete loc no and wayleave yes  *********************************************

carriageway_d_t_concrete_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution & Trunk') ]

carriageway_d_t_concrete_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_d_t_concrete_way = carriageway_d_t_concrete_way['length'].sum()

num_of_rows_cw_d_t_concrete_2x96_way = carriageway_d_t_concrete_2x96_way['length'].sum()

num_of_rows_cw_d_t_concrete_total_way = num_of_rows_cw_d_t_concrete_way + num_of_rows_cw_d_t_concrete_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K39"] = num_of_rows_cw_d_t_concrete_total_way
wb.save(filename)

# *********************** carriageway_D+T_unmade and grassverge loc no and wayleave yes  *********************************************

carriageway_d_t_grassverge_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution & Trunk')]

carriageway_d_t_unmade_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution & Trunk')]

carriageway_d_t_grassverge_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

carriageway_d_t_unmade_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]


num_of_rows_cw_d_t_grassverge_way = carriageway_d_t_grassverge_way['length'].sum()

num_of_rows_cw_d_t_unmade_way = carriageway_d_t_unmade_way['length'].sum()

num_of_rows_cw_d_t_grassverge_2x96_way = carriageway_d_t_grassverge_2x96_way['length'].sum()

num_of_rows_cw_d_t_unmade_2x96_way = carriageway_d_t_unmade_2x96_way['length'].sum()

num_of_rows_cw_d_t_unmade_grassverge_total_way = num_of_rows_cw_d_t_grassverge_way + num_of_rows_cw_d_t_unmade_way + num_of_rows_cw_d_t_grassverge_2x96_way + num_of_rows_cw_d_t_unmade_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L39"] = num_of_rows_cw_d_t_unmade_grassverge_total_way
wb.save(filename)

# *********************** carriageway_D+T_tarmac loc no and wayleave yes  *********************************************

carriageway_d_t_tarmac_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Distribution & Trunk') ]

carriageway_d_t_tarmac_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_d_t_tarmac_way = carriageway_d_t_tarmac_way['length'].sum()

num_of_rows_cw_d_t_tarmac_2x96_way = carriageway_d_t_tarmac_2x96_way['length'].sum()

num_of_rows_cw_d_t_tarmac_total_way = num_of_rows_cw_d_t_tarmac_way + num_of_rows_cw_d_t_tarmac_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M39"] = num_of_rows_cw_d_t_tarmac_total_way
wb.save(filename)

# *********************** carriageway_D_modular loc no and wayleave yes  *********************************************

carriageway_d_modular_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular' ) & (df['type'] == 'Distribution') ]

carriageway_d_modular_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_d_modular_way = carriageway_d_modular_way['length'].sum()

num_of_rows_cw_d_modular_2x96_way = carriageway_d_modular_2x96_way['length'].sum()

num_of_rows_cw_d_modular_total_way = num_of_rows_cw_d_modular_way + num_of_rows_cw_d_modular_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J40"] = num_of_rows_cw_d_modular_total_way
wb.save(filename)

# *********************** carriageway_D_concrete loc no and wayleave yes  *********************************************

carriageway_d_concrete_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution') ]

carriageway_d_concrete_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_d_concrete_way = carriageway_d_concrete_way['length'].sum()

num_of_rows_cw_d_concrete_2x96_way = carriageway_d_concrete_2x96_way['length'].sum()

num_of_rows_cw_d_concrete_total_way = num_of_rows_cw_d_concrete_way + num_of_rows_cw_d_concrete_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K40"] = num_of_rows_cw_d_concrete_total_way
wb.save(filename)

# *********************** carriageway_D_unmade and grassverge loc no and wayleave yes  *********************************************

carriageway_d_grassverge_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution')]

carriageway_d_unmade_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution')]

carriageway_d_grassverge_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

carriageway_d_unmade_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]


num_of_rows_cw_d_grassverge_way = carriageway_d_grassverge_way['length'].sum()

num_of_rows_cw_d_unmade_way = carriageway_d_unmade_way['length'].sum()

num_of_rows_cw_d_grassverge_2x96_way = carriageway_d_grassverge_2x96_way['length'].sum()

num_of_rows_cw_d_unmade_2x96_way = carriageway_d_unmade_2x96_way['length'].sum()

num_of_rows_cw_d_unmade_grassverge_total_way = num_of_rows_cw_d_grassverge_way + num_of_rows_cw_d_unmade_way + num_of_rows_cw_d_grassverge_2x96_way + num_of_rows_cw_d_unmade_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L40"] = num_of_rows_cw_d_unmade_grassverge_total_way
wb.save(filename)

# *********************** carriageway_D_tarmac loc no and wayleave yes  *********************************************

carriageway_d_tarmac_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Distribution') ]

carriageway_d_tarmac_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_d_tarmac_way = carriageway_d_tarmac_way['length'].sum()

num_of_rows_cw_d_tarmac_2x96_way = carriageway_d_tarmac_2x96_way['length'].sum()

num_of_rows_cw_d_tarmac_total_way = num_of_rows_cw_d_tarmac_way + num_of_rows_cw_d_tarmac_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M40"] = num_of_rows_cw_d_tarmac_total_way
wb.save(filename)

# *********************** carriageway_S+D+T_modular loc no and wayleave yes  *********************************************

carriageway_s_d_t_modular_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular' ) & (df['type'] == 'Acces, Distribution & Trunk') ]

carriageway_s_d_t_modular_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_d_t_modular_way = carriageway_s_d_t_modular_way['length'].sum()

num_of_rows_cw_s_d_t_modular_2x96_way = carriageway_s_d_t_modular_2x96_way['length'].sum()

num_of_rows_cw_s_d_t_modular_total_way = num_of_rows_cw_s_d_t_modular_way + num_of_rows_cw_s_d_t_modular_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J41"] = num_of_rows_cw_s_d_t_modular_total_way
wb.save(filename)

# *********************** carriageway_S+D+T_concrete loc no and wayleave yes  *********************************************

carriageway_s_d_t_concrete_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Acces, Distribution & Trunk') ]

carriageway_s_d_t_concrete_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_d_t_concrete_way = carriageway_s_d_t_concrete_way['length'].sum()

num_of_rows_cw_s_d_t_concrete_2x96_way = carriageway_s_d_t_concrete_2x96_way['length'].sum()

num_of_rows_cw_s_d_t_concrete_total_way = num_of_rows_cw_s_d_t_concrete_way + num_of_rows_cw_s_d_t_concrete_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K41"] = num_of_rows_cw_s_d_t_concrete_total_way
wb.save(filename)

# *********************** carriageway_S+D+T_unmade and grassverge loc no and wayleave yes  *********************************************

carriageway_s_d_t_grassverge_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Acces, Distribution & Trunk')]

carriageway_s_d_t_unmade_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Acces, Distribution & Trunk')]

carriageway_s_d_t_grassverge_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

carriageway_s_d_t_unmade_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]


num_of_rows_cw_s_d_t_grassverge_way = carriageway_s_d_t_grassverge_way['length'].sum()

num_of_rows_cw_s_d_t_unmade_way = carriageway_s_d_t_unmade_way['length'].sum()

num_of_rows_cw_s_d_t_grassverge_2x96_way = carriageway_s_d_t_grassverge_2x96_way['length'].sum()

num_of_rows_cw_s_d_t_unmade_2x96_way = carriageway_s_d_t_unmade_2x96_way['length'].sum()

num_of_rows_cw_s_d_t_unmade_grassverge_total_way = num_of_rows_cw_s_d_t_grassverge_way + num_of_rows_cw_s_d_t_unmade_way + num_of_rows_cw_s_d_t_grassverge_2x96_way + num_of_rows_cw_s_d_t_unmade_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L41"] = num_of_rows_cw_s_d_t_unmade_grassverge_total_way
wb.save(filename)

# *********************** carriageway_S+D+T_tarmac loc no and wayleave yes  *********************************************

carriageway_s_d_t_tarmac_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Acces, Distribution & Trunk') ]

carriageway_s_d_t_tarmac_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_d_t_tarmac_way = carriageway_s_d_t_tarmac_way['length'].sum()

num_of_rows_cw_s_d_t_tarmac_2x96_way = carriageway_s_d_t_tarmac_2x96_way['length'].sum()

num_of_rows_cw_s_d_t_tarmac_total_way = num_of_rows_cw_s_d_t_tarmac_way + num_of_rows_cw_s_d_t_tarmac_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M41"] = num_of_rows_cw_s_d_t_tarmac_total_way
wb.save(filename)

# *********************** carriageway_S+D_modular loc no and wayleave yes  *********************************************

carriageway_s_d_modular_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular' ) & (df['type'] == 'Access & Distribution') ]

carriageway_s_d_modular_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_d_modular_way = carriageway_s_d_modular_way['length'].sum()

num_of_rows_cw_s_d_modular_2x96_way = carriageway_s_d_modular_2x96_way['length'].sum()

num_of_rows_cw_s_d_modular_total_way = num_of_rows_cw_s_d_modular_way + num_of_rows_cw_s_d_modular_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J42"] = num_of_rows_cw_s_d_modular_total_way
wb.save(filename)

# *********************** carriageway_S+D_concrete loc no and wayleave yes  *********************************************

carriageway_s_d_concrete_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Distribution') ]

carriageway_s_d_concrete_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_d_concrete_way = carriageway_s_d_concrete_way['length'].sum()

num_of_rows_cw_s_d_concrete_2x96_way = carriageway_s_d_concrete_2x96_way['length'].sum()

num_of_rows_cw_s_d_concrete_total_way = num_of_rows_cw_s_d_concrete_way + num_of_rows_cw_s_d_concrete_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K42"] = num_of_rows_cw_s_d_concrete_total_way
wb.save(filename)

# *********************** carriageway_S+D_unmade and grassverge loc no and wayleave yes  *********************************************

carriageway_s_d_grassverge_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Distribution')]

carriageway_s_d_unmade_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Distribution')]

carriageway_s_d_grassverge_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

carriageway_s_d_unmade_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]


num_of_rows_cw_s_d_grassverge_way = carriageway_s_d_grassverge_way['length'].sum()

num_of_rows_cw_s_d_unmade_way = carriageway_s_d_unmade_way['length'].sum()

num_of_rows_cw_s_d_grassverge_2x96_way = carriageway_s_d_grassverge_2x96_way['length'].sum()

num_of_rows_cw_s_d_unmade_2x96_way = carriageway_s_d_unmade_2x96_way['length'].sum()

num_of_rows_cw_s_d_unmade_grassverge_total_way = num_of_rows_cw_s_d_grassverge_way + num_of_rows_cw_s_d_unmade_way + num_of_rows_cw_s_d_grassverge_2x96_way + num_of_rows_cw_s_d_unmade_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L42"] = num_of_rows_cw_s_d_unmade_grassverge_total_way
wb.save(filename)

# *********************** carriageway_S+D_tarmac loc no and wayleave yes  *********************************************

carriageway_s_d_tarmac_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access & Distribution') ]

carriageway_s_d_tarmac_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_d_tarmac_way = carriageway_s_d_tarmac_way['length'].sum()

num_of_rows_cw_s_d_tarmac_2x96_way = carriageway_s_d_tarmac_2x96_way['length'].sum()

num_of_rows_cw_s_d_tarmac_total_way = num_of_rows_cw_s_d_tarmac_way + num_of_rows_cw_s_d_tarmac_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M42"] = num_of_rows_cw_s_d_tarmac_total_way
wb.save(filename)

# *********************** carriageway_S_modular loc no and wayleave yes  *********************************************

carriageway_s_modular_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular' ) & (df['type'] == 'Access') ]

carriageway_s_modular_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_modular_way = carriageway_s_modular_way['length'].sum()

num_of_rows_cw_s_modular_2x96_way = carriageway_s_modular_2x96_way['length'].sum()

num_of_rows_cw_s_modular_total_way = num_of_rows_cw_s_modular_way + num_of_rows_cw_s_modular_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J43"] = num_of_rows_cw_s_modular_total_way
wb.save(filename)

# *********************** carriageway_S_concrete loc no and wayleave yes  *********************************************

carriageway_s_concrete_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Access') ]

carriageway_s_concrete_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_concrete_way = carriageway_s_concrete_way['length'].sum()

num_of_rows_cw_s_concrete_2x96_way = carriageway_s_concrete_2x96_way['length'].sum()

num_of_rows_cw_s_concrete_total_way = num_of_rows_cw_s_concrete_way + num_of_rows_cw_s_concrete_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K43"] = num_of_rows_cw_s_concrete_total_way
wb.save(filename)

# *********************** carriageway_S_unmade and grassverge loc no and wayleave yes  *********************************************

carriageway_s_grassverge_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access')]

carriageway_s_unmade_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Access')]

carriageway_s_grassverge_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

carriageway_s_unmade_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]


num_of_rows_cw_s_grassverge_way = carriageway_s_grassverge_way['length'].sum()

num_of_rows_cw_s_unmade_way = carriageway_s_unmade_way['length'].sum()

num_of_rows_cw_s_grassverge_2x96_way = carriageway_s_grassverge_2x96_way['length'].sum()

num_of_rows_cw_s_unmade_2x96_way = carriageway_s_unmade_2x96_way['length'].sum()

num_of_rows_cw_s_unmade_grassverge_total_way = num_of_rows_cw_s_grassverge_way + num_of_rows_cw_s_unmade_way + num_of_rows_cw_s_grassverge_2x96_way + num_of_rows_cw_s_unmade_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L43"] = num_of_rows_cw_s_unmade_grassverge_total_way
wb.save(filename)

# *********************** carriageway_S_tarmac loc no and wayleave yes  *********************************************

carriageway_s_tarmac_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access') ]

carriageway_s_tarmac_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_s_tarmac_way = carriageway_s_tarmac_way['length'].sum()

num_of_rows_cw_s_tarmac_2x96_way = carriageway_s_tarmac_2x96_way['length'].sum()

num_of_rows_cw_s_tarmac_total_way = num_of_rows_cw_s_tarmac_way + num_of_rows_cw_s_tarmac_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M43"] = num_of_rows_cw_s_tarmac_total_way
wb.save(filename)

# *********************** carriageway_T_modular loc no and wayleave yes  *********************************************

carriageway_t_modular_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular' ) & (df['type'] == 'Trunk') ]

carriageway_t_modular_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_t_modular_way = carriageway_t_modular_way['length'].sum()

num_of_rows_cw_t_modular_2x96_way = carriageway_t_modular_2x96_way['length'].sum()

num_of_rows_cw_t_modular_total_way = num_of_rows_cw_t_modular_way + num_of_rows_cw_t_modular_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J44"] = num_of_rows_cw_t_modular_total_way
wb.save(filename)

# *********************** carriageway_T_concrete loc no and wayleave yes  *********************************************

carriageway_t_concrete_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Trunk') ]

carriageway_t_concrete_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_t_concrete_way = carriageway_t_concrete_way['length'].sum()

num_of_rows_cw_t_concrete_2x96_way = carriageway_t_concrete_2x96_way['length'].sum()

num_of_rows_cw_t_concrete_total_way = num_of_rows_cw_t_concrete_way + num_of_rows_cw_t_concrete_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K44"] = num_of_rows_cw_t_concrete_total_way
wb.save(filename)

# *********************** carriageway_t_unmade and grassverge loc no and wayleave yes  *********************************************

carriageway_t_grassverge_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Trunk')]

carriageway_t_unmade_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Trunk')]

carriageway_t_grassverge_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

carriageway_t_unmade_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]


num_of_rows_cw_t_grassverge_way = carriageway_t_grassverge_way['length'].sum()

num_of_rows_cw_t_unmade_way = carriageway_t_unmade_way['length'].sum()

num_of_rows_cw_t_grassverge_2x96_way = carriageway_t_grassverge_2x96_way['length'].sum()

num_of_rows_cw_t_unmade_2x96_way = carriageway_t_unmade_2x96_way['length'].sum()

num_of_rows_cw_t_unmade_grassverge_total_way = num_of_rows_cw_t_grassverge_way + num_of_rows_cw_t_unmade_way + num_of_rows_cw_t_grassverge_2x96_way + num_of_rows_cw_t_unmade_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L44"] = num_of_rows_cw_t_unmade_grassverge_total_way
wb.save(filename)

# *********************** carriageway_T_tarmac loc no and wayleave yes  *********************************************

carriageway_t_tarmac_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Trunk') ]

carriageway_t_tarmac_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Carriageway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_cw_t_tarmac_way = carriageway_t_tarmac_way['length'].sum()

num_of_rows_cw_t_tarmac_2x96_way = carriageway_t_tarmac_2x96_way['length'].sum()

num_of_rows_cw_t_tarmac_total_way = num_of_rows_cw_t_tarmac_way + num_of_rows_cw_t_tarmac_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M44"] = num_of_rows_cw_t_tarmac_total_way
wb.save(filename)

# *********************** footway_S+T_modular loc no and wayleave yes  *********************************************

footway_s_t_modular_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular' ) & (df['type'] == 'Access & Trunk') ]

footway_s_t_modular_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_t_modular_way = footway_s_t_modular_way['length'].sum()

num_of_rows_fw_s_t_modular_2x96_way = footway_s_t_modular_2x96_way['length'].sum()

num_of_rows_fw_s_t_modular_total_way = num_of_rows_fw_s_t_modular_way + num_of_rows_fw_s_t_modular_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J49"] = num_of_rows_fw_s_t_modular_total_way
wb.save(filename)

# *********************** footway_S+T_concrete loc no and wayleave yes  *********************************************

footway_s_t_concrete_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Trunk') ]

footway_s_t_concrete_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_t_concrete_way = footway_s_t_concrete_way['length'].sum()

num_of_rows_fw_s_t_concrete_2x96_way = footway_s_t_concrete_2x96_way['length'].sum()

num_of_rows_fw_s_t_concrete_total_way = num_of_rows_fw_s_t_concrete_way + num_of_rows_fw_s_t_concrete_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K49"] = num_of_rows_fw_s_t_concrete_total_way
wb.save(filename)

# *********************** footway_S+T_unmade and grassverge loc no and wayleave yes  *********************************************

footway_s_t_grassverge_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Trunk')]

footway_s_t_unmade_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Trunk')]

footway_s_t_grassverge_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Trunk') & (df['96mm'] == '2x96')]

footway_s_t_unmade_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Trunk') & (df['96mm'] == '2x96')]


num_of_rows_fw_s_t_grassverge_way = footway_s_t_grassverge_way['length'].sum()

num_of_rows_fw_s_t_unmade_way = footway_s_t_unmade_way['length'].sum()

num_of_rows_fw_s_t_grassverge_2x96_way = footway_s_t_grassverge_2x96_way['length'].sum()

num_of_rows_fw_s_t_unmade_2x96_way = footway_s_t_unmade_2x96_way['length'].sum()

num_of_rows_fw_s_t_unmade_grassverge_total_way = num_of_rows_fw_s_t_grassverge_way + num_of_rows_fw_s_t_unmade_way + num_of_rows_fw_s_t_grassverge_2x96_way + num_of_rows_fw_s_t_unmade_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L49"] = num_of_rows_fw_s_t_unmade_grassverge_total_way
wb.save(filename)

# *********************** footway_S+T_tarmac loc no and wayleave yes  *********************************************

footway_s_t_tarmac_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access & Trunk') ]

footway_s_t_tarmac_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac') & (df['type'] == 'Access & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_t_tarmac_way = footway_s_t_tarmac_way['length'].sum()

num_of_rows_fw_s_t_tarmac_2x96_way = footway_s_t_tarmac_2x96_way['length'].sum()

num_of_rows_fw_s_t_tarmac_total_way = num_of_rows_fw_s_t_tarmac_way + num_of_rows_fw_s_t_tarmac_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M49"] = num_of_rows_fw_s_t_tarmac_total_way
wb.save(filename)

# *********************** footway_D+T_modular loc no and wayleave yes  *********************************************

footway_d_t_modular_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular' ) & (df['type'] == 'Distribution & Trunk') ]

footway_d_t_modular_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_d_t_modular_way = footway_d_t_modular_way['length'].sum()

num_of_rows_fw_d_t_modular_2x96_way = footway_d_t_modular_2x96_way['length'].sum()

num_of_rows_fw_d_t_modular_total_way = num_of_rows_fw_d_t_modular_way + num_of_rows_fw_d_t_modular_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J50"] = num_of_rows_fw_d_t_modular_total_way
wb.save(filename)

# *********************** footway_D+T_concrete loc no and wayleave yes  *********************************************

footway_d_t_concrete_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution & Trunk') ]

footway_d_t_concrete_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_d_t_concrete_way = footway_d_t_concrete_way['length'].sum()

num_of_rows_fw_d_t_concrete_2x96_way = footway_d_t_concrete_2x96_way['length'].sum()

num_of_rows_fw_d_t_concrete_total_way = num_of_rows_fw_d_t_concrete_way + num_of_rows_fw_d_t_concrete_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K50"] = num_of_rows_fw_d_t_concrete_total_way
wb.save(filename)

# *********************** footway_D+T_unmade and grassverge loc no and wayleave yes  *********************************************

footway_d_t_grassverge_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution & Trunk')]

footway_d_t_unmade_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution & Trunk')]

footway_d_t_grassverge_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution & Trunk') & (df['96mm'] == '2x96')]

footway_d_t_unmade_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution & Trunk') & (df['96mm'] == '2x96')]


num_of_rows_fw_d_t_grassverge_way = footway_d_t_grassverge_way['length'].sum()

num_of_rows_fw_d_t_unmade_way = footway_d_t_unmade_way['length'].sum()

num_of_rows_fw_d_t_grassverge_2x96_way = footway_d_t_grassverge_2x96_way['length'].sum()

num_of_rows_fw_d_t_unmade_2x96_way = footway_d_t_unmade_2x96_way['length'].sum()

num_of_rows_fw_d_t_unmade_grassverge_total_way = num_of_rows_fw_d_t_grassverge_way + num_of_rows_fw_d_t_unmade_way + num_of_rows_fw_d_t_grassverge_2x96_way + num_of_rows_fw_d_t_unmade_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L50"] = num_of_rows_fw_d_t_unmade_grassverge_total_way
wb.save(filename)

# *********************** footway_D+T_tarmac loc no and wayleave yes  *********************************************

footway_d_t_tarmac_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Distribution & Trunk') ]

footway_d_t_tarmac_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac') & (df['type'] == 'Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_d_t_tarmac_way = footway_d_t_tarmac_way['length'].sum()

num_of_rows_fw_d_t_tarmac_2x96_way = footway_d_t_tarmac_2x96_way['length'].sum()

num_of_rows_fw_d_t_tarmac_total_way = num_of_rows_fw_d_t_tarmac_way + num_of_rows_fw_d_t_tarmac_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M50"] = num_of_rows_fw_d_t_tarmac_total_way
wb.save(filename)

# *********************** footway_D_modular loc no and wayleave yes  *********************************************

footway_d_modular_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular' ) & (df['type'] == 'Distribution') ]

footway_d_modular_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_d_modular_way = footway_d_modular_way['length'].sum()

num_of_rows_fw_d_modular_2x96_way = footway_d_modular_2x96_way['length'].sum()

num_of_rows_fw_d_modular_total_way = num_of_rows_fw_d_modular_way + num_of_rows_fw_d_modular_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J51"] = num_of_rows_fw_d_modular_total_way
wb.save(filename)


# *********************** footway_D_concrete loc no and wayleave yes  *********************************************

footway_d_concrete_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution') ]

footway_d_concrete_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_d_concrete_way = footway_d_concrete_way['length'].sum()

num_of_rows_fw_d_concrete_2x96_way = footway_d_concrete_2x96_way['length'].sum()

num_of_rows_fw_d_concrete_total_way = num_of_rows_fw_d_concrete_way + num_of_rows_fw_d_concrete_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K51"] = num_of_rows_fw_d_concrete_total_way
wb.save(filename)

# *********************** footway_D_unmade and grassverge loc no and wayleave yes  *********************************************

footway_d_grassverge_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution')]

footway_d_unmade_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution')]

footway_d_grassverge_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Distribution') & (df['96mm'] == '2x96')]

footway_d_unmade_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Distribution') & (df['96mm'] == '2x96')]


num_of_rows_fw_d_grassverge_way = footway_d_grassverge_way['length'].sum()

num_of_rows_fw_d_unmade_way = footway_d_unmade_way['length'].sum()

num_of_rows_fw_d_grassverge_2x96_way = footway_d_grassverge_2x96_way['length'].sum()

num_of_rows_fw_d_unmade_2x96_way = footway_d_unmade_2x96_way['length'].sum()

num_of_rows_fw_d_unmade_grassverge_total_way = num_of_rows_fw_d_grassverge_way + num_of_rows_fw_d_unmade_way + num_of_rows_fw_d_grassverge_2x96_way + num_of_rows_fw_d_unmade_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L51"] = num_of_rows_fw_d_unmade_grassverge_total_way
wb.save(filename)

# *********************** footway_D_tarmac loc no and wayleave yes  *********************************************

footway_d_tarmac_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Distribution') ]

footway_d_tarmac_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac') & (df['type'] == 'Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_d_tarmac_way = footway_d_tarmac_way['length'].sum()

num_of_rows_fw_d_tarmac_2x96_way = footway_d_tarmac_2x96_way['length'].sum()

num_of_rows_fw_d_tarmac_total_way = num_of_rows_fw_d_tarmac_way + num_of_rows_fw_d_tarmac_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M51"] = num_of_rows_fw_d_tarmac_total_way
wb.save(filename)

# *********************** footway_S_D_T_modular loc no and wayleave yes  *********************************************

footway_s_d_t_modular_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular' ) & (df['type'] == 'Acces, Distribution & Trunk') ]

footway_s_d_t_modular_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_d_t_modular_way = footway_s_d_t_modular_way['length'].sum()

num_of_rows_fw_s_d_t_modular_2x96_way = footway_s_d_t_modular_2x96_way['length'].sum()

num_of_rows_fw_s_d_t_modular_total_way = num_of_rows_fw_s_d_t_modular_way + num_of_rows_fw_s_d_t_modular_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J52"] = num_of_rows_fw_s_d_t_modular_total_way
wb.save(filename)


# *********************** footway_S+D+T_concrete loc no and wayleave yes  *********************************************

footway_s_d_t_concrete_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Acces, Distribution & Trunk') ]

footway_s_d_t_concrete_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_d_t_concrete_way = footway_s_d_t_concrete_way['length'].sum()

num_of_rows_fw_s_d_t_concrete_2x96_way = footway_s_d_t_concrete_2x96_way['length'].sum()

num_of_rows_fw_s_d_t_concrete_total_way = num_of_rows_fw_s_d_t_concrete_way + num_of_rows_fw_s_d_t_concrete_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K52"] = num_of_rows_fw_s_d_t_concrete_total_way
wb.save(filename)

# *********************** footway_S+D+T_unmade and grassverge loc no and wayleave yes  *********************************************

footway_s_d_t_grassverge_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Acces, Distribution & Trunk')]

footway_s_d_t_unmade_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Acces, Distribution & Trunk')]

footway_s_d_t_grassverge_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Acces, Distribution & Trunk') & (df['96mm'] == '2x96')]

footway_s_d_t_unmade_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Acces, Distribution & Trunk') & (df['96mm'] == '2x96')]


num_of_rows_fw_s_d_t_grassverge_way = footway_s_d_t_grassverge_way['length'].sum()

num_of_rows_fw_s_d_t_unmade_way = footway_s_d_t_unmade_way['length'].sum()

num_of_rows_fw_s_d_t_grassverge_2x96_way = footway_s_d_t_grassverge_2x96_way['length'].sum()

num_of_rows_fw_s_d_t_unmade_2x96_way = footway_s_d_t_unmade_2x96_way['length'].sum()

num_of_rows_fw_s_d_t_unmade_grassverge_total_way = num_of_rows_fw_s_d_t_grassverge_way + num_of_rows_fw_s_d_t_unmade_way + num_of_rows_fw_s_d_t_grassverge_2x96_way + num_of_rows_fw_s_d_t_unmade_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L52"] = num_of_rows_fw_s_d_t_unmade_grassverge_total_way
wb.save(filename)

# *********************** footway_S+D+T_tarmac loc no and wayleave yes  *********************************************

footway_s_d_t_tarmac_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Acces, Distribution & Trunk') ]

footway_s_d_t_tarmac_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac') & (df['type'] == 'Acces, Distribution & Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_d_t_tarmac_way = footway_s_d_t_tarmac_way['length'].sum()

num_of_rows_fw_s_d_t_tarmac_2x96_way = footway_s_d_t_tarmac_2x96_way['length'].sum()

num_of_rows_fw_s_d_t_tarmac_total_way = num_of_rows_fw_s_d_t_tarmac_way + num_of_rows_fw_s_d_t_tarmac_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M52"] = num_of_rows_fw_s_d_t_tarmac_total_way
wb.save(filename)

# *********************** footway_S+D_modular loc no and wayleave yes  *********************************************

footway_s_d_modular_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular' ) & (df['type'] == 'Access & Distribution') ]

footway_s_d_modular_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Modular') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_d_modular_way = footway_s_d_modular_way['length'].sum()

num_of_rows_fw_s_d_modular_2x96_way = footway_s_d_modular_2x96_way['length'].sum()

num_of_rows_fw_s_d_modular_total_way = num_of_rows_fw_s_d_modular_way + num_of_rows_fw_s_d_modular_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["J53"] = num_of_rows_fw_s_d_modular_total_way
wb.save(filename)


# *********************** footway_S+D_concrete loc no and wayleave yes  *********************************************

footway_s_d_concrete_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Distribution') ]

footway_s_d_concrete_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Concrete') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_d_concrete_way = footway_s_d_concrete_way['length'].sum()

num_of_rows_fw_s_d_concrete_2x96_way = footway_s_d_concrete_2x96_way['length'].sum()

num_of_rows_fw_s_d_concrete_total_way = num_of_rows_fw_s_d_concrete_way + num_of_rows_fw_s_d_concrete_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["K53"] = num_of_rows_fw_s_d_concrete_total_way
wb.save(filename)

# *********************** footway_S+D_unmade and grassverge loc no and wayleave yes  *********************************************

footway_s_d_grassverge_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Distribution')]

footway_s_d_unmade_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Distribution')]

footway_s_d_grassverge_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access & Distribution') & (df['96mm'] == '2x96')]

footway_s_d_unmade_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Unmade') & (df['type'] == 'Access & Distribution') & (df['96mm'] == '2x96')]


num_of_rows_fw_s_d_grassverge_way = footway_s_d_grassverge_way['length'].sum()

num_of_rows_fw_s_d_unmade_way = footway_s_d_unmade_way['length'].sum()

num_of_rows_fw_s_d_grassverge_2x96_way = footway_s_d_grassverge_2x96_way['length'].sum()

num_of_rows_fw_s_d_unmade_2x96_way = footway_s_d_unmade_2x96_way['length'].sum()

num_of_rows_fw_s_d_unmade_grassverge_total_way = num_of_rows_fw_s_d_grassverge_way + num_of_rows_fw_s_d_unmade_way + num_of_rows_fw_s_d_grassverge_2x96_way + num_of_rows_fw_s_d_unmade_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["L53"] = num_of_rows_fw_s_d_unmade_grassverge_total_way
wb.save(filename)

# *********************** footway_S+D_tarmac loc no and wayleave yes  *********************************************

footway_s_d_tarmac_way= df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac' ) & (df['type'] == 'Access & Distribution') ]

footway_s_d_tarmac_2x96_way = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == True) & (df['material'] == 'Tarmac') & (df['type'] == 'Access & Distribution')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_d_tarmac_way = footway_s_d_tarmac_way['length'].sum()

num_of_rows_fw_s_d_tarmac_2x96_way = footway_s_d_tarmac_2x96_way['length'].sum()

num_of_rows_fw_s_d_tarmac_total_way = num_of_rows_fw_s_d_tarmac_way + num_of_rows_fw_s_d_tarmac_2x96_way


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["M53"] = num_of_rows_fw_s_d_tarmac_total_way
wb.save(filename)

# *********************** footway_S_modular loc no and wayleave no  *********************************************

footway_s_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Access') ]

footway_s_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_modular = footway_s_modular['length'].sum()

num_of_rows_fw_s_modular_2x96 = footway_s_modular_2x96['length'].sum()

num_of_rows_fw_s_modular_total = num_of_rows_fw_s_modular + num_of_rows_fw_s_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B54"] = num_of_rows_fw_s_modular_total
wb.save(filename)


# *********************** footway_S_concrete loc no and wayleave no  *********************************************

footway_s_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access') ]

footway_s_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Access')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_s_concrete = footway_s_concrete['length'].sum()

num_of_rows_fw_s_concrete_2x96 = footway_s_concrete_2x96['length'].sum()

num_of_rows_fw_s_concrete_total = num_of_rows_fw_s_concrete + num_of_rows_fw_s_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C54"] = num_of_rows_fw_s_concrete_total
wb.save(filename)

# *********************** footway_S_unmade and grassverge loc no and wayleave no  *********************************************

footway_s_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access')]

footway_s_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access')]

footway_s_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Access') & (df['96mm'] == '2x96')]

footway_s_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Access') & (df['96mm'] == '2x96')]


num_of_rows_fw_s_grassverge = footway_s_grassverge['length'].sum()

num_of_rows_fw_s_unmade = footway_s_unmade['length'].sum()

num_of_rows_fw_s_grassverge_2x96 = footway_s_grassverge_2x96['length'].sum()

num_of_rows_fw_s_unmade_2x96 = footway_s_unmade_2x96['length'].sum()

num_of_rows_fw_s_unmade_grassverge_total = num_of_rows_fw_s_grassverge + num_of_rows_fw_s_unmade + num_of_rows_fw_s_grassverge_2x96 + num_of_rows_fw_s_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D54"] = num_of_rows_fw_s_unmade_grassverge_total
wb.save(filename)

# *********************** footway_S_tarmac loc no and wayleave no  ************************************************

footway_s_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Access') ]

footway_s_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Access')
                         &  (df['96mm'] == '2x96') ]

num_of_rows_fw_s_tarmac = footway_s_tarmac['length'].sum()

num_of_rows_fw_s_tarmac_2x96 = footway_s_tarmac_2x96['length'].sum()

num_of_rows_fw_s_tarmac_total = num_of_rows_fw_s_tarmac + num_of_rows_fw_s_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E54"] = num_of_rows_fw_s_tarmac_total
wb.save(filename)

# *********************** footway_T_modular loc no and wayleave no  *********************************************

footway_t_modular = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular' ) & (df['type'] == 'Trunk') ]

footway_t_modular_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Modular') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_t_modular = footway_t_modular['length'].sum()

num_of_rows_fw_t_modular_2x96 = footway_t_modular_2x96['length'].sum()

num_of_rows_fw_t_modular_total = num_of_rows_fw_t_modular + num_of_rows_fw_t_modular_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["B55"] = num_of_rows_fw_t_modular_total
wb.save(filename)


# *********************** footway_t_concrete loc no and wayleave no  *********************************************

footway_t_concrete = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Trunk') ]

footway_t_concrete_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Concrete') & (df['type'] == 'Trunk')
                         & (df['96mm'] == '2x96')]

num_of_rows_fw_t_concrete = footway_t_concrete['length'].sum()

num_of_rows_fw_t_concrete_2x96 = footway_t_concrete_2x96['length'].sum()

num_of_rows_fw_t_concrete_total = num_of_rows_fw_t_concrete + num_of_rows_fw_t_concrete_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["C55"] = num_of_rows_fw_t_concrete_total
wb.save(filename)

# *********************** footway_T_unmade and grassverge loc no and wayleave no  *********************************************

footway_t_grassverge = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Trunk')]

footway_t_unmade = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Trunk')]

footway_t_grassverge_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Grass Verge') & (df['type'] == 'Trunk') & (df['96mm'] == '2x96')]

footway_t_unmade_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Unmade') & (df['type'] == 'Trunk') & (df['96mm'] == '2x96')]


num_of_rows_fw_t_grassverge = footway_t_grassverge['length'].sum()

num_of_rows_fw_t_unmade = footway_t_unmade['length'].sum()

num_of_rows_fw_t_grassverge_2x96 = footway_t_grassverge_2x96['length'].sum()

num_of_rows_fw_t_unmade_2x96 = footway_t_unmade_2x96['length'].sum()

num_of_rows_fw_t_unmade_grassverge_total = num_of_rows_fw_t_grassverge + num_of_rows_fw_t_unmade + num_of_rows_fw_t_grassverge_2x96 + num_of_rows_fw_t_unmade_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["D55"] = num_of_rows_fw_t_unmade_grassverge_total
wb.save(filename)

# *********************** footway_T_tarmac loc no and wayleave no  ************************************************

footway_t_tarmac = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Trunk') ]

footway_t_tarmac_2x96 = df.loc[(df['state'] == 'Planned') & (df['surface'] == 'Footway') & (df['loc'] == False)
                         & (df['wayleave'] == False) & (df['material'] == 'Tarmac') & (df['type'] == 'Trunk')
                         &  (df['96mm'] == '2x96') ]

num_of_rows_fw_t_tarmac = footway_t_tarmac['length'].sum()

num_of_rows_fw_t_tarmac_2x96 = footway_t_tarmac_2x96['length'].sum()

num_of_rows_fw_t_tarmac_total = num_of_rows_fw_t_tarmac + num_of_rows_fw_t_tarmac_2x96


filename = "Documents/Output/BOQ and BOM.xlsx"

n = 4
wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[n]]
ws_tables = []
ws["E55"] = num_of_rows_fw_t_tarmac_total
wb.save(filename)